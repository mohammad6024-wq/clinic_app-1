from PySide6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, 
                               QGridLayout, QScrollArea, QPushButton,
                               QMessageBox, QDialog, QTableWidget, QTableWidgetItem,
                               QHeaderView, QLineEdit, QComboBox, QGroupBox,
                               QCompleter, QApplication, QFrame, QFileDialog)
from PySide6.QtCore import Qt, QTimer
from database.engine import SessionLocal
from database.models import Appointment, Doctor, Patient, User
import jdatetime
import re
import os
import tempfile
import webbrowser

# ==================== نرمالایز کردن متن برای جستجو ====================
def normalize_text(text):
    """تبدیل ی و ک عربی به فارسی و حذف تفاوت‌های املایی"""
    if not text:
        return ""
    text = str(text)
    text = text.replace('ي', 'ی')
    text = text.replace('ك', 'ک')
    text = text.replace('إ', 'ا').replace('أ', 'ا').replace('آ', 'ا')
    text = re.sub(r'[ًٌٍَُّْ]', '', text)
    return text.lower()

# ==================== خروجی اکسل حرفه‌ای ====================
def export_to_excel(data, headers, title, parent=None):
    """خروجی اکسل حرفه‌ای با openpyxl"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = title[:31] if title else "Report"
        
        # تنظیم راست‌چین
        ws.sheet_view.rightToLeft = True
        
        # استایل‌ها
        header_font = Font(name='Tahoma', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1e40af', end_color='1e40af', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        cell_font = Font(name='Tahoma', size=10)
        cell_alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        
        total_font = Font(name='Tahoma', size=11, bold=True, color='FFFFFF')
        total_fill = PatternFill(start_color='2563eb', end_color='2563eb', fill_type='solid')
        
        thin_border = Border(
            left=Side(style='thin', color='cbd5e1'),
            right=Side(style='thin', color='cbd5e1'),
            top=Side(style='thin', color='cbd5e1'),
            bottom=Side(style='thin', color='cbd5e1')
        )
        
        # هدرها
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # داده‌ها
        totals = {}
        numeric_columns = []
        
        for col_idx, header in enumerate(headers):
            if any(word in header for word in ["مبلغ", "هزینه", "قیمت", "تعداد", "شناسه", "مبلغ نهایی"]):
                numeric_columns.append(col_idx)
                totals[col_idx] = 0
        
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data):
                cell = ws.cell(row=row_idx, column=col_idx + 1, value=value)
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = thin_border
                
                if col_idx in numeric_columns:
                    try:
                        clean_value = str(value).replace(',', '').replace('تومان', '').strip()
                        if clean_value:
                            num_val = float(clean_value)
                            totals[col_idx] = totals.get(col_idx, 0) + num_val
                    except:
                        pass
        
        # عرض ستون‌ها
        for col_idx, header in enumerate(headers, 1):
            max_length = len(header)
            for row_idx in range(2, len(data) + 2):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            adjusted_width = min(max_length + 3, 40)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
        
        # ردیف جمع
        if totals:
            total_row = len(data) + 2
            for col_idx in range(len(headers)):
                if col_idx in totals:
                    total_value = int(totals[col_idx])
                    formatted_total = f"{total_value:,}"
                    cell = ws.cell(row=total_row, column=col_idx + 1, value=formatted_total)
                elif col_idx == 0:
                    cell = ws.cell(row=total_row, column=1, value="جمع کل")
                else:
                    cell = ws.cell(row=total_row, column=col_idx + 1, value="-")
                
                cell.font = total_font
                cell.fill = total_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
        
        # فیلتر خودکار
        max_row = len(data) + 2 if totals else len(data) + 1
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max_row}"
        
        # ارتفاع ردیف‌ها
        ws.row_dimensions[1].height = 25
        for row_idx in range(2, max_row + 1):
            ws.row_dimensions[row_idx].height = 20
        
        # فریز کردن هدر
        ws.freeze_panes = 'A2'
        
        # ذخیره
        filepath, _ = QFileDialog.getSaveFileName(parent, "ذخیره فایل اکسل", 
                                                   f"{title}.xlsx", 
                                                   "Excel Files (*.xlsx)")
        if filepath:
            wb.save(filepath)
            QMessageBox.information(parent, "موفق", 
                f"✅ فایل اکسل با موفقیت ذخیره شد!\n\n📁 {filepath}\n📊 {len(data)} رکورد")
            return True
            
    except ImportError:
        msg = QMessageBox(parent)
        msg.setWindowTitle("کتابخانه missing")
        msg.setText("❌ برای خروجی اکسل حرفه‌ای، لطفاً اجرا کنید:\n\npip install openpyxl")
        msg.exec()
        return False
    except Exception as e:
        QMessageBox.critical(parent, "خطا", f"خطا در ذخیره اکسل:\n{str(e)}")
        return False
    return False

# ==================== خروجی PDF ====================
def export_to_pdf(data, headers, title, parent=None):
    """خروجی PDF حرفه‌ای با HTML"""
    try:
        html_content = f"""
        <!DOCTYPE html>
        <html dir="rtl">
        <head>
            <meta charset="utf-8">
            <title>{title}</title>
            <style>
                @page {{
                    size: A4 landscape;
                    margin: 15mm;
                }}
                body {{
                    font-family: 'Tahoma', 'Segoe UI', sans-serif;
                    direction: rtl;
                    margin: 0;
                    padding: 20px;
                    background: #f8fafc;
                }}
                .container {{
                    max-width: 100%;
                    background: white;
                    border-radius: 16px;
                    box-shadow: 0 4px 20px rgba(0,0,0,0.1);
                    overflow: hidden;
                }}
                .header {{
                    background: linear-gradient(135deg, #1e3a5f, #2563eb);
                    color: white;
                    padding: 20px;
                    text-align: center;
                }}
                .header h1 {{
                    margin: 0;
                    font-size: 24px;
                }}
                .header p {{
                    margin: 5px 0 0;
                    opacity: 0.9;
                    font-size: 12px;
                }}
                .info {{
                    padding: 15px 20px;
                    background: #f1f5f9;
                    border-bottom: 1px solid #e2e8f0;
                    display: flex;
                    justify-content: space-between;
                    font-size: 12px;
                }}
                table {{
                    width: 100%;
                    border-collapse: collapse;
                    font-size: 11px;
                }}
                th {{
                    background: linear-gradient(135deg, #2563eb, #1e40af);
                    color: white;
                    padding: 12px 8px;
                    font-weight: bold;
                    text-align: center;
                    border: 1px solid #3b82f6;
                }}
                td {{
                    padding: 8px;
                    text-align: center;
                    border: 1px solid #cbd5e1;
                }}
                tr:nth-child(even) {{
                    background-color: #f8fafc;
                }}
                tr:hover {{
                    background-color: #e0f2fe;
                }}
                .footer {{
                    padding: 15px 20px;
                    background: #f1f5f9;
                    border-top: 1px solid #e2e8f0;
                    display: flex;
                    justify-content: space-between;
                    font-size: 11px;
                    color: #475569;
                }}
                .total-row {{
                    background: #dbeafe;
                    font-weight: bold;
                }}
                .total-row td {{
                    background: #dbeafe;
                    font-weight: bold;
                }}
                @media print {{
                    body {{
                        background: white;
                        padding: 0;
                    }}
                    .container {{
                        box-shadow: none;
                    }}
                    .no-print {{
                        display: none;
                    }}
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>📊 {title}</h1>
                    <p>تاریخ چاپ: {jdatetime.datetime.now().strftime("%Y/%m/%d - %H:%M")}</p>
                </div>
                <div class="info">
                    <span>📋 تعداد کل رکوردها: {len(data)}</span>
                    <span>🏢 مرکز مشاوره آرامش</span>
                </div>
                <table>
                    <thead>
                        <tr>
        """
        
        for h in headers:
            html_content += f"<th>{h}</th>"
        
        html_content += """
                        </tr>
                    </thead>
                    <tbody>
        """
        
        totals = {}
        numeric_columns = []
        
        for col_idx, header in enumerate(headers):
            if any(word in header for word in ["مبلغ", "هزینه", "قیمت", "تعداد"]):
                numeric_columns.append(col_idx)
                totals[col_idx] = 0
        
        for row in data:
            html_content += "<tr>"
            for col_idx, cell in enumerate(row):
                if col_idx in numeric_columns:
                    try:
                        val = str(cell).replace(",", "").replace("تومان", "").strip()
                        if val.isdigit():
                            totals[col_idx] = totals.get(col_idx, 0) + int(val)
                    except:
                        pass
                html_content += f"<td>{cell}</td>"
            html_content += "</tr>"
        
        if totals:
            html_content += '<tr class="total-row">'
            for col_idx, header in enumerate(headers):
                if col_idx in totals:
                    html_content += f"<td><strong>جمع: {totals[col_idx]:,}</strong></td>"
                elif col_idx == 0:
                    html_content += '<td><strong>جمع کل</strong></td>'
                else:
                    html_content += "<td></td>"
            html_content += "</tr>"
        
        html_content += """
                    </tbody>
                </table>
                <div class="footer">
                    <span>🔹 این گزارش توسط سیستم مدیریت مرکز مشاوره آرامش تولید شده است</span>
                    <span>✅ معتبر و رسمی</span>
                </div>
            </div>
            <div class="no-print" style="text-align: center; margin-top: 20px;">
                <button onclick="window.print()" style="background: #3b82f6; color: white; border: none; padding: 10px 20px; border-radius: 8px; cursor: pointer;">
                    🖨️ چاپ / ذخیره PDF
                </button>
                <button onclick="window.close()" style="background: #ef4444; color: white; border: none; padding: 10px 20px; border-radius: 8px; cursor: pointer; margin-right: 10px;">
                    ❌ بستن
                </button>
            </div>
            <script>
                setTimeout(function() { window.print(); }, 500);
            </script>
        </body>
        </html>
        """
        
        temp_file = os.path.join(tempfile.gettempdir(), f"{title.replace(' ', '_')}.html")
        with open(temp_file, 'w', encoding='utf-8') as f:
            f.write(html_content)
        webbrowser.open(temp_file)
        return True
    except Exception as e:
        QMessageBox.critical(parent, "خطا", f"خطا در ایجاد PDF:\n{str(e)}")
        return False

# ==================== دیالوگ نمایش لیست ====================
class ListViewDialog(QDialog):
    def __init__(self, title, items, headers, appointments_obj, parent=None):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setModal(True)
        self.resize(1300, 650)
        self.setLayoutDirection(Qt.RightToLeft)
        self.all_appointments = appointments_obj
        self.session = SessionLocal()
        self.current_title = title
        self.current_headers = headers
        
        layout = QVBoxLayout(self)
        
        title_lbl = QLabel(title)
        title_lbl.setStyleSheet("font-size: 16px; font-weight: bold; color: #2563eb; padding: 10px; background: #eff6ff; border-radius: 10px;")
        title_lbl.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_lbl)
        
        filter_frame = QFrame()
        filter_frame.setStyleSheet("QFrame { background: #f8fafc; border-radius: 10px; padding: 10px; }")
        filter_layout = QGridLayout(filter_frame)
        filter_layout.setSpacing(10)
        
        filter_layout.addWidget(QLabel("استاد:"), 0, 0)
        self.doctor_filter = QComboBox()
        self.doctor_filter.setEditable(True)
        self.doctor_filter.setMinimumWidth(220)
        self.doctor_filter.addItem("همه اساتید")
        
        doctors = self.session.query(Doctor).order_by(Doctor.name).all()
        self.doctor_names = ["همه اساتید"]
        for doc in doctors:
            self.doctor_filter.addItem(doc.name)
            self.doctor_names.append(doc.name)
        
        self.doctor_filter.setStyleSheet("padding: 6px; border: 1px solid #cbd5e1; border-radius: 8px;")
        completer = QCompleter(self.doctor_names[1:])
        completer.setCaseSensitivity(Qt.CaseInsensitive)
        completer.setFilterMode(Qt.MatchContains)
        self.doctor_filter.setCompleter(completer)
        filter_layout.addWidget(self.doctor_filter, 0, 1)
        
        filter_layout.addWidget(QLabel("وضعیت:"), 0, 2)
        self.status_filter = QComboBox()
        self.status_filter.addItems(["همه", "فعال", "انجام شده", "کنسل استاد", "کنسل مراجع"])
        self.status_filter.setStyleSheet("padding: 6px; border: 1px solid #cbd5e1; border-radius: 8px;")
        filter_layout.addWidget(self.status_filter, 0, 3)
        
        filter_layout.addWidget(QLabel("نوع نوبت:"), 1, 0)
        self.type_filter = QComboBox()
        self.type_filter.addItems(["همه", "حضوری", "تلفنی", "آنلاین"])
        self.type_filter.setStyleSheet("padding: 6px; border: 1px solid #cbd5e1; border-radius: 8px;")
        filter_layout.addWidget(self.type_filter, 1, 1)
        
        filter_layout.addWidget(QLabel("جستجوی متن:"), 1, 2)
        self.search_box = QLineEdit()
        self.search_box.setPlaceholderText("جستجو در نام مراجع یا نام استاد...")
        self.search_box.setStyleSheet("padding: 6px; border: 1px solid #cbd5e1; border-radius: 8px;")
        self.search_box.textChanged.connect(self.apply_filter)
        filter_layout.addWidget(self.search_box, 1, 3)
        
        btn_layout = QHBoxLayout()
        
        self.export_excel_btn = QPushButton("📎 خروجی اکسل حرفه‌ای")
        self.export_excel_btn.setStyleSheet("background: #10b981; color: white; padding: 8px 15px; border-radius: 8px; font-weight: bold;")
        self.export_excel_btn.clicked.connect(self.export_excel)
        btn_layout.addWidget(self.export_excel_btn)
        
        self.export_pdf_btn = QPushButton("📄 خروجی PDF")
        self.export_pdf_btn.setStyleSheet("background: #8b5cf6; color: white; padding: 8px 15px; border-radius: 8px; font-weight: bold;")
        self.export_pdf_btn.clicked.connect(self.export_pdf)
        btn_layout.addWidget(self.export_pdf_btn)
        
        self.reset_btn = QPushButton("🔄 بازنشانی")
        self.reset_btn.setStyleSheet("background: #6b7280; color: white; padding: 8px 15px; border-radius: 8px;")
        btn_layout.addWidget(self.reset_btn)
        
        btn_layout.addStretch()
        filter_layout.addLayout(btn_layout, 2, 0, 1, 4)
        
        layout.addWidget(filter_frame)
        
        self.table = QTableWidget()
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setAlternatingRowColors(True)
        layout.addWidget(self.table)
        
        self.stats_lbl = QLabel("")
        self.stats_lbl.setStyleSheet("font-size: 12px; color: #6b7280; padding: 5px;")
        layout.addWidget(self.stats_lbl)
        
        close_btn = QPushButton("بستن")
        close_btn.setStyleSheet("background: #ef4444; color: white; padding: 8px 20px; border-radius: 8px;")
        close_btn.clicked.connect(self.accept)
        layout.addWidget(close_btn, alignment=Qt.AlignCenter)
        
        self.all_items = items
        self.load_data(items)
        
        self.doctor_filter.currentTextChanged.connect(self.apply_filter)
        self.status_filter.currentTextChanged.connect(self.apply_filter)
        self.type_filter.currentTextChanged.connect(self.apply_filter)
    
    def load_data(self, items):
        self.table.setRowCount(len(items))
        for row, item in enumerate(items):
            for col, val in enumerate(item):
                self.table.setItem(row, col, QTableWidgetItem(str(val)))
        self.stats_lbl.setText(f"جمع کل: {len(items)} مورد")
        self.current_displayed_data = items
    
    def apply_filter(self):
        doctor = self.doctor_filter.currentText()
        status = self.status_filter.currentText()
        app_type = self.type_filter.currentText()
        search = self.search_box.text().lower()
        search_normalized = normalize_text(search)
        
        filtered = []
        for app in self.all_appointments:
            if doctor != "همه اساتید" and hasattr(app, 'doctor') and app.doctor != doctor:
                continue
            if status != "همه" and hasattr(app, 'status') and app.status != status:
                continue
            if app_type != "همه" and hasattr(app, 'type') and app.type != app_type:
                continue
            if search:
                patient_name = normalize_text(getattr(app, 'patient_name', "") or "")
                doctor_name = normalize_text(getattr(app, 'doctor', "") or "")
                if search_normalized not in patient_name and search_normalized not in doctor_name:
                    continue
            filtered.append(app)
        
        data = []
        for app in filtered:
            row = []
            for col in range(self.table.columnCount()):
                header = self.table.horizontalHeaderItem(col).text()
                if header == "شناسه":
                    row.append(app.id)
                elif header == "تاریخ":
                    row.append(getattr(app, 'date', "") or "")
                elif header == "ساعت":
                    row.append(getattr(app, 'time', "") or "")
                elif header == "استاد":
                    row.append(getattr(app, 'doctor', "") or "")
                elif header == "نام مراجع":
                    row.append(getattr(app, 'patient_name', "") or "")
                elif header == "تلفن":
                    row.append(getattr(app, 'phone', "") or "")
                elif header == "وضعیت":
                    row.append(getattr(app, 'status', "") or "")
                elif header == "نوع":
                    row.append(getattr(app, 'type', "") or "")
                elif header == "مبلغ":
                    row.append(f"{int(getattr(app, 'final_cost', 0) or 0):,}")
                else:
                    row.append("")
            data.append(row)
        
        self.load_data(data)
    
    def reset_filter(self):
        self.doctor_filter.setCurrentIndex(0)
        self.status_filter.setCurrentIndex(0)
        self.type_filter.setCurrentIndex(0)
        self.search_box.clear()
        self.load_data(self.all_items)
    
    def export_excel(self):
        headers = [self.table.horizontalHeaderItem(i).text() for i in range(self.table.columnCount())]
        data = []
        for row in range(self.table.rowCount()):
            row_data = []
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                row_data.append(item.text() if item else "")
            data.append(row_data)
        export_to_excel(data, headers, self.current_title, self)
    
    def export_pdf(self):
        headers = [self.table.horizontalHeaderItem(i).text() for i in range(self.table.columnCount())]
        data = []
        for row in range(self.table.rowCount()):
            row_data = []
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                row_data.append(item.text() if item else "")
            data.append(row_data)
        export_to_pdf(data, headers, self.current_title, self)
    
    def __del__(self):
        if hasattr(self, 'session'):
            try:
                self.session.close()
            except:
                pass

# ==================== تقویم شمسی ====================
class PersianCalendarDialog(QDialog):
    def __init__(self, parent=None, callback=None):
        super().__init__(parent)
        self.setWindowTitle("انتخاب تاریخ شمسی")
        self.setModal(True)
        self.setFixedSize(550, 520)
        self.callback = callback
        self.current_date = jdatetime.date.today()
        
        layout = QVBoxLayout(self)
        
        header = QLabel("📅 تقویم شمسی")
        header.setStyleSheet("font-size: 18px; font-weight: bold; color: #2563eb; padding: 10px; background: #eff6ff; border-radius: 10px;")
        header.setAlignment(Qt.AlignCenter)
        layout.addWidget(header)
        
        nav_layout = QHBoxLayout()
        
        self.prev_year_btn = QPushButton("<<")
        self.prev_year_btn.setFixedSize(45, 35)
        self.prev_year_btn.setStyleSheet("background: #e2e8f0; border-radius: 8px;")
        self.prev_year_btn.clicked.connect(self.prev_year)
        nav_layout.addWidget(self.prev_year_btn)
        
        self.prev_month_btn = QPushButton("<")
        self.prev_month_btn.setFixedSize(45, 35)
        self.prev_month_btn.setStyleSheet("background: #e2e8f0; border-radius: 8px;")
        self.prev_month_btn.clicked.connect(self.prev_month)
        nav_layout.addWidget(self.prev_month_btn)
        
        nav_layout.addStretch()
        
        self.year_combo = QComboBox()
        self.year_combo.setStyleSheet("padding: 5px; border-radius: 8px; min-width: 80px;")
        for y in range(1390, jdatetime.date.today().year + 10):
            self.year_combo.addItem(str(y))
        self.year_combo.setCurrentText(str(self.current_date.year))
        self.year_combo.currentTextChanged.connect(self.update_calendar)
        nav_layout.addWidget(self.year_combo)
        
        self.month_combo = QComboBox()
        self.month_combo.setStyleSheet("padding: 5px; border-radius: 8px; min-width: 100px;")
        months = ["فروردین", "اردیبهشت", "خرداد", "تیر", "مرداد", "شهریور", 
                  "مهر", "آبان", "آذر", "دی", "بهمن", "اسفند"]
        self.month_combo.addItems(months)
        self.month_combo.setCurrentIndex(self.current_date.month - 1)
        self.month_combo.currentTextChanged.connect(self.update_calendar)
        nav_layout.addWidget(self.month_combo)
        
        nav_layout.addStretch()
        
        self.next_month_btn = QPushButton(">")
        self.next_month_btn.setFixedSize(45, 35)
        self.next_month_btn.setStyleSheet("background: #e2e8f0; border-radius: 8px;")
        self.next_month_btn.clicked.connect(self.next_month)
        nav_layout.addWidget(self.next_month_btn)
        
        self.next_year_btn = QPushButton(">>")
        self.next_year_btn.setFixedSize(45, 35)
        self.next_year_btn.setStyleSheet("background: #e2e8f0; border-radius: 8px;")
        self.next_year_btn.clicked.connect(self.next_year)
        nav_layout.addWidget(self.next_year_btn)
        
        layout.addLayout(nav_layout)
        
        week_layout = QHBoxLayout()
        week_days = ["شنبه", "یکشنبه", "دوشنبه", "سه‌شنبه", "چهارشنبه", "پنج‌شنبه", "جمعه"]
        for day in week_days:
            lbl = QLabel(day)
            lbl.setAlignment(Qt.AlignCenter)
            lbl.setStyleSheet("font-weight: bold; color: #2563eb; padding: 10px; background: #dbeafe; border-radius: 8px;")
            week_layout.addWidget(lbl)
        layout.addLayout(week_layout)
        
        self.day_grid = QGridLayout()
        self.day_grid.setSpacing(8)
        layout.addLayout(self.day_grid)
        
        today_btn = QPushButton("📅 امروز")
        today_btn.setStyleSheet("background: #10b981; color: white; padding: 10px; border-radius: 10px; font-weight: bold;")
        today_btn.clicked.connect(self.go_today)
        layout.addWidget(today_btn, alignment=Qt.AlignCenter)
        
        self.update_calendar()
    
    def prev_year(self):
        self.current_date = jdatetime.date(self.current_date.year - 1, self.current_date.month, 1)
        self.year_combo.setCurrentText(str(self.current_date.year))
    
    def next_year(self):
        self.current_date = jdatetime.date(self.current_date.year + 1, self.current_date.month, 1)
        self.year_combo.setCurrentText(str(self.current_date.year))
    
    def prev_month(self):
        if self.current_date.month == 1:
            self.current_date = jdatetime.date(self.current_date.year - 1, 12, 1)
        else:
            self.current_date = jdatetime.date(self.current_date.year, self.current_date.month - 1, 1)
        self.year_combo.setCurrentText(str(self.current_date.year))
        self.month_combo.setCurrentIndex(self.current_date.month - 1)
    
    def next_month(self):
        if self.current_date.month == 12:
            self.current_date = jdatetime.date(self.current_date.year + 1, 1, 1)
        else:
            self.current_date = jdatetime.date(self.current_date.year, self.current_date.month + 1, 1)
        self.year_combo.setCurrentText(str(self.current_date.year))
        self.month_combo.setCurrentIndex(self.current_date.month - 1)
    
    def go_today(self):
        self.current_date = jdatetime.date.today()
        self.year_combo.setCurrentText(str(self.current_date.year))
        self.month_combo.setCurrentIndex(self.current_date.month - 1)
    
    def update_calendar(self):
        for i in reversed(range(self.day_grid.count())):
            w = self.day_grid.itemAt(i).widget()
            if w:
                w.deleteLater()
        
        year = int(self.year_combo.currentText())
        month = self.month_combo.currentIndex() + 1
        
        first_day = jdatetime.date(year, month, 1)
        start_weekday = first_day.weekday()
        
        if month <= 6:
            days = 31
        elif month <= 11:
            days = 30
        else:
            days = 30 if first_day.isleap() else 29
        
        row, col = 0, start_weekday
        for day in range(1, days + 1):
            date = jdatetime.date(year, month, day)
            btn = QPushButton(str(day))
            btn.setFixedSize(55, 45)
            
            if date == jdatetime.date.today():
                btn.setStyleSheet("""
                    QPushButton {
                        background: #3b82f6;
                        color: white;
                        border-radius: 12px;
                        font-size: 14px;
                        font-weight: bold;
                    }
                """)
            else:
                btn.setStyleSheet("""
                    QPushButton {
                        background: #f1f5f9;
                        color: #1e293b;
                        border-radius: 12px;
                        font-size: 14px;
                    }
                    QPushButton:hover {
                        background: #cbd5e1;
                    }
                """)
            
            btn.clicked.connect(lambda checked, d=date: self.select_date(d))
            self.day_grid.addWidget(btn, row, col)
            col += 1
            if col > 6:
                col = 0
                row += 1
    
    def select_date(self, date):
        if self.callback:
            self.callback(date.strftime("%Y/%m/%d"))
        self.accept()

# ==================== فیلتر پیشرفته ====================
class AdvancedFilterDialog(QDialog):
    def __init__(self, parent=None, session=None):
        super().__init__(parent)
        self.parent = parent
        self.session = session
        self.setWindowTitle("فیلتر پیشرفته نوبت‌ها")
        self.setModal(True)
        self.resize(600, 550)
        self.setLayoutDirection(Qt.RightToLeft)
        
        layout = QVBoxLayout(self)
        
        title = QLabel("🔍 فیلتر پیشرفته نوبت‌ها")
        title.setStyleSheet("font-size: 18px; font-weight: bold; color: #2563eb; padding: 10px;")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
        
        form = QGroupBox("معیارهای جستجو")
        form_layout = QGridLayout(form)
        form_layout.setSpacing(15)
        
        form_layout.addWidget(QLabel("از تاریخ:"), 0, 0)
        start_layout = QHBoxLayout()
        self.start_edit = QLineEdit()
        self.start_edit.setPlaceholderText("انتخاب تاریخ...")
        self.start_edit.setReadOnly(True)
        self.start_edit.setStyleSheet("padding: 8px; border-radius: 8px; border: 1px solid #cbd5e1;")
        start_btn = QPushButton("📅")
        start_btn.setFixedWidth(40)
        start_btn.setStyleSheet("background: #e2e8f0; border-radius: 8px;")
        start_btn.clicked.connect(lambda: self.show_calendar(self.start_edit))
        start_layout.addWidget(self.start_edit)
        start_layout.addWidget(start_btn)
        default_start = jdatetime.date.today() - jdatetime.timedelta(days=30)
        self.start_edit.setText(default_start.strftime("%Y/%m/%d"))
        form_layout.addLayout(start_layout, 0, 1)
        
        form_layout.addWidget(QLabel("تا تاریخ:"), 1, 0)
        end_layout = QHBoxLayout()
        self.end_edit = QLineEdit()
        self.end_edit.setPlaceholderText("انتخاب تاریخ...")
        self.end_edit.setReadOnly(True)
        self.end_edit.setStyleSheet("padding: 8px; border-radius: 8px; border: 1px solid #cbd5e1;")
        end_btn = QPushButton("📅")
        end_btn.setFixedWidth(40)
        end_btn.setStyleSheet("background: #e2e8f0; border-radius: 8px;")
        end_btn.clicked.connect(lambda: self.show_calendar(self.end_edit))
        end_layout.addWidget(self.end_edit)
        end_layout.addWidget(end_btn)
        self.end_edit.setText(jdatetime.date.today().strftime("%Y/%m/%d"))
        form_layout.addLayout(end_layout, 1, 1)
        
        form_layout.addWidget(QLabel("استاد:"), 2, 0)
        self.doctor_combo = QComboBox()
        self.doctor_combo.setEditable(True)
        self.doctor_combo.setMinimumWidth(250)
        self.doctor_combo.addItem("همه اساتید")
        
        doctors = self.session.query(Doctor).order_by(Doctor.name).all()
        doctor_names = ["همه اساتید"]
        for doc in doctors:
            self.doctor_combo.addItem(doc.name)
            doctor_names.append(doc.name)
        
        self.doctor_combo.setStyleSheet("padding: 8px; border: 1px solid #cbd5e1; border-radius: 8px;")
        completer = QCompleter(doctor_names[1:])
        completer.setCaseSensitivity(Qt.CaseInsensitive)
        completer.setFilterMode(Qt.MatchContains)
        self.doctor_combo.setCompleter(completer)
        form_layout.addWidget(self.doctor_combo, 2, 1)
        
        form_layout.addWidget(QLabel("وضعیت:"), 3, 0)
        self.status_combo = QComboBox()
        self.status_combo.addItems(["همه", "فعال", "انجام شده", "کنسل استاد", "کنسل مراجع"])
        self.status_combo.setStyleSheet("padding: 8px; border: 1px solid #cbd5e1; border-radius: 8px;")
        form_layout.addWidget(self.status_combo, 3, 1)
        
        form_layout.addWidget(QLabel("نوع نوبت:"), 4, 0)
        self.type_combo = QComboBox()
        self.type_combo.addItems(["همه", "حضوری", "تلفنی", "آنلاین"])
        self.type_combo.setStyleSheet("padding: 8px; border: 1px solid #cbd5e1; border-radius: 8px;")
        form_layout.addWidget(self.type_combo, 4, 1)
        
        layout.addWidget(form)
        
        btn_layout = QHBoxLayout()
        search_btn = QPushButton("🔍 جستجو")
        search_btn.setStyleSheet("background: #3b82f6; color: white; padding: 10px 20px; border-radius: 10px; font-size: 14px;")
        search_btn.clicked.connect(self.search)
        btn_layout.addWidget(search_btn)
        
        cancel_btn = QPushButton("لغو")
        cancel_btn.setStyleSheet("background: #ef4444; color: white; padding: 10px 20px; border-radius: 10px; font-size: 14px;")
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)
        layout.addLayout(btn_layout)
    
    def show_calendar(self, target):
        def on_date_selected(date_str):
            target.setText(date_str)
        dialog = PersianCalendarDialog(self, on_date_selected)
        dialog.exec()
    
    def search(self):
        start = self.start_edit.text()
        end = self.end_edit.text()
        
        if not start or not end:
            QMessageBox.warning(self, "خطا", "لطفاً هر دو تاریخ را انتخاب کنید")
            return
        
        try:
            start_date = jdatetime.datetime.strptime(start, "%Y/%m/%d").date()
            end_date = jdatetime.datetime.strptime(end, "%Y/%m/%d").date()
        except:
            QMessageBox.warning(self, "خطا", "تاریخ نامعتبر است")
            return
        
        query = self.session.query(Appointment).filter(
            Appointment.date >= start_date.strftime("%Y/%m/%d"),
            Appointment.date <= end_date.strftime("%Y/%m/%d")
        )
        
        doctor = self.doctor_combo.currentText()
        if doctor != "همه اساتید":
            query = query.filter(Appointment.doctor == doctor)
        
        status = self.status_combo.currentText()
        if status != "همه":
            query = query.filter(Appointment.status == status)
        
        app_type = self.type_combo.currentText()
        if app_type != "همه":
            query = query.filter(Appointment.type == app_type)
        
        apps = query.all()
        self.accept()
        
        if not apps:
            QMessageBox.information(self.parent, "اطلاع", "هیچ نوبتی با معیارهای انتخاب شده یافت نشد")
            return
        
        data = [[a.id, a.date, a.time, a.doctor, a.patient_name, a.phone, a.type, a.status, f"{int(a.final_cost or 0):,}"] for a in apps]
        headers = ["شناسه", "تاریخ", "ساعت", "استاد", "نام مراجع", "تلفن", "نوع", "وضعیت", "مبلغ"]
        dialog = ListViewDialog(f"نتایج جستجو ({len(apps)} مورد)", data, headers, apps, self.parent)
        dialog.exec()

# ==================== داشبورد اصلی ====================
class StatsDashboard(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent = parent
        self.session = SessionLocal()
        self.setup_ui()
        self.load_stats()
        self.timer = QTimer()
        self.timer.timeout.connect(self.load_stats)
        self.timer.start(60000)
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        
        header = QLabel("📊 داشبورد آماری مرکز مشاوره")
        header.setStyleSheet("font-size: 20px; font-weight: bold; color: #1e293b; padding: 15px; background: #f1f5f9; border-radius: 15px;")
        header.setAlignment(Qt.AlignCenter)
        layout.addWidget(header)
        
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea { border: none; background: transparent; }")
        
        self.cards_widget = QWidget()
        self.cards_layout = QGridLayout(self.cards_widget)
        self.cards_layout.setSpacing(15)
        
        self.cards = {}
        
        cards_info = [
            ("نوبت‌های امروز", "today", "📅", "#3b82f6"),
            ("نوبت‌های هفته جاری", "week", "📆", "#10b981"),
            ("نوبت‌های ماه جاری", "month", "📅", "#8b5cf6"),
            ("نوبت‌های انجام شده", "done", "✅", "#059669"),
            ("نوبت‌های لغو شده", "cancelled", "❌", "#dc2626"),
            ("نوبت‌های ارجاعی", "referred", "🔄", "#f59e0b"),
            ("نوبت‌های زوجی", "couple", "💑", "#ec4899"),
            ("لیست اساتید", "doctors", "👨‍⚕️", "#06b6d4"),
            ("لیست مراجعین", "patients", "👥", "#84cc16"),
            ("کاربران فعال", "users", "👤", "#a855f7"),
            ("فیلتر پیشرفته", "filter", "🔍", "#ef4444"),
        ]
        
        row, col = 0, 0
        for title, key, icon, color in cards_info:
            card = self.create_card(title, icon, color)
            card.clicked.connect(lambda checked, k=key: self.on_click(k))
            self.cards[key] = card
            self.cards_layout.addWidget(card, row, col)
            col += 1
            if col >= 3:
                col = 0
                row += 1
        
        scroll.setWidget(self.cards_widget)
        layout.addWidget(scroll)
        
        refresh_btn = QPushButton("🔄 بروزرسانی")
        refresh_btn.setStyleSheet("background: #3b82f6; color: white; padding: 10px 25px; border-radius: 12px; font-weight: bold;")
        refresh_btn.clicked.connect(self.load_stats)
        
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        btn_layout.addWidget(refresh_btn)
        btn_layout.addStretch()
        layout.addLayout(btn_layout)
    
    def create_card(self, title, icon, color):
        card = QPushButton()
        card.setFixedSize(210, 110)
        card.setCursor(Qt.PointingHandCursor)
        card.setStyleSheet(f"""
            QPushButton {{
                background-color: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 {color}, stop:1 {color}cc);
                border-radius: 20px;
                border: none;
            }}
            QPushButton:hover {{
                background-color: {color};
            }}
        """)
        
        layout = QVBoxLayout(card)
        layout.setContentsMargins(15, 10, 15, 10)
        
        top = QHBoxLayout()
        icon_lbl = QLabel(icon)
        icon_lbl.setStyleSheet("font-size: 30px; background: transparent;")
        title_lbl = QLabel(title)
        title_lbl.setStyleSheet("font-size: 12px; color: white; background: transparent; font-weight: bold;")
        top.addWidget(icon_lbl)
        top.addWidget(title_lbl)
        top.addStretch()
        layout.addLayout(top)
        
        value_lbl = QLabel("0")
        value_lbl.setStyleSheet("font-size: 36px; font-weight: bold; color: white; background: transparent;")
        value_lbl.setAlignment(Qt.AlignCenter)
        layout.addWidget(value_lbl)
        
        card.value_label = value_lbl
        return card
    
    def on_click(self, key):
        if key == "today":
            today = jdatetime.date.today().strftime("%Y/%m/%d")
            apps = self.session.query(Appointment).filter(Appointment.date == today).all()
            if apps:
                data = [[a.id, a.time, a.doctor, a.patient_name, a.phone, a.type, a.status] for a in apps]
                headers = ["شناسه", "ساعت", "استاد", "نام مراجع", "تلفن", "نوع", "وضعیت"]
                ListViewDialog(f"نوبت‌های امروز ({len(apps)} مورد)", data, headers, apps, self).exec()
            else:
                QMessageBox.information(self, "اطلاع", "هیچ نوبتی برای امروز وجود ندارد")
        elif key == "week":
            today = jdatetime.date.today()
            week_start = today - jdatetime.timedelta(days=today.weekday())
            week_end = week_start + jdatetime.timedelta(days=6)
            apps = self.session.query(Appointment).filter(
                Appointment.date >= week_start.strftime("%Y/%m/%d"),
                Appointment.date <= week_end.strftime("%Y/%m/%d")
            ).all()
            if apps:
                data = [[a.id, a.date, a.time, a.doctor, a.patient_name, a.type, a.status] for a in apps]
                headers = ["شناسه", "تاریخ", "ساعت", "استاد", "نام مراجع", "نوع", "وضعیت"]
                ListViewDialog(f"نوبت‌های هفته جاری ({len(apps)} مورد)", data, headers, apps, self).exec()
            else:
                QMessageBox.information(self, "اطلاع", "هیچ نوبتی در این هفته وجود ندارد")
        elif key == "month":
            today = jdatetime.date.today()
            month_start = jdatetime.date(today.year, today.month, 1)
            if today.month == 12:
                month_end = jdatetime.date(today.year + 1, 1, 1) - jdatetime.timedelta(days=1)
            else:
                month_end = jdatetime.date(today.year, today.month + 1, 1) - jdatetime.timedelta(days=1)
            apps = self.session.query(Appointment).filter(
                Appointment.date >= month_start.strftime("%Y/%m/%d"),
                Appointment.date <= month_end.strftime("%Y/%m/%d")
            ).all()
            if apps:
                month_names = ["فروردین", "اردیبهشت", "خرداد", "تیر", "مرداد", "شهریور", "مهر", "آبان", "آذر", "دی", "بهمن", "اسفند"]
                data = [[a.id, a.date, a.time, a.doctor, a.patient_name, a.type, a.status] for a in apps]
                headers = ["شناسه", "تاریخ", "ساعت", "استاد", "نام مراجع", "نوع", "وضعیت"]
                ListViewDialog(f"نوبت‌های ماه {month_names[today.month-1]} ({len(apps)} مورد)", data, headers, apps, self).exec()
            else:
                QMessageBox.information(self, "اطلاع", "هیچ نوبتی در این ماه وجود ندارد")
        elif key == "done":
            apps = self.session.query(Appointment).filter(Appointment.status == "انجام شده").all()
            if apps:
                data = [[a.id, a.date, a.time, a.doctor, a.patient_name] for a in apps]
                headers = ["شناسه", "تاریخ", "ساعت", "استاد", "نام مراجع"]
                ListViewDialog(f"نوبت‌های انجام شده ({len(apps)} مورد)", data, headers, apps, self).exec()
            else:
                QMessageBox.information(self, "اطلاع", "هیچ نوبت انجام شده‌ای وجود ندارد")
        elif key == "cancelled":
            apps = self.session.query(Appointment).filter(Appointment.status.in_(["کنسل استاد", "کنسل مراجع"])).all()
            if apps:
                data = [[a.id, a.date, a.time, a.doctor, a.patient_name, a.status] for a in apps]
                headers = ["شناسه", "تاریخ", "ساعت", "استاد", "نام مراجع", "وضعیت"]
                ListViewDialog(f"نوبت‌های لغو شده ({len(apps)} مورد)", data, headers, apps, self).exec()
            else:
                QMessageBox.information(self, "اطلاع", "هیچ نوبت لغو شده‌ای وجود ندارد")
        elif key == "referred":
            apps = self.session.query(Appointment).filter(Appointment.ref_type != "آزاد").all()
            if apps:
                data = [[a.id, a.date, a.doctor, a.patient_name, a.ref_type] for a in apps]
                headers = ["شناسه", "تاریخ", "استاد", "نام مراجع", "نوع ارجاع"]
                ListViewDialog(f"نوبت‌های ارجاعی ({len(apps)} مورد)", data, headers, apps, self).exec()
            else:
                QMessageBox.information(self, "اطلاع", "هیچ نوبت ارجاعی وجود ندارد")
        elif key == "couple":
            apps = self.session.query(Appointment).filter(Appointment.patient2_name != "").all()
            if apps:
                data = [[a.id, a.date, a.doctor, a.patient_name, a.patient2_name] for a in apps]
                headers = ["شناسه", "تاریخ", "استاد", "نام مراجع", "نام همسر"]
                ListViewDialog(f"نوبت‌های زوجی ({len(apps)} مورد)", data, headers, apps, self).exec()
            else:
                QMessageBox.information(self, "اطلاع", "هیچ نوبت زوجی وجود ندارد")
        elif key == "doctors":
            doctors = self.session.query(Doctor).all()
            data = [[d.id, d.name, d.spec or "-", d.phone or "-"] for d in doctors]
            headers = ["شناسه", "نام", "تخصص", "تلفن"]
            ListViewDialog(f"لیست اساتید ({len(doctors)} نفر)", data, headers, doctors, self).exec()
        elif key == "patients":
            patients = self.session.query(Patient).all()
            data = [[p.id, p.name, p.nat_id or "-", p.phone or "-"] for p in patients]
            headers = ["شناسه", "نام", "کد ملی", "تلفن"]
            ListViewDialog(f"لیست مراجعین ({len(patients)} نفر)", data, headers, patients, self).exec()
        elif key == "users":
            users = self.session.query(User).filter(User.is_active == 1).all()
            data = [[u.id, u.username, u.name or "-", u.role] for u in users]
            headers = ["شناسه", "نام کاربری", "نام کامل", "نقش"]
            ListViewDialog(f"کاربران فعال ({len(users)} نفر)", data, headers, users, self).exec()
        elif key == "filter":
            AdvancedFilterDialog(self, self.session).exec()
    
    def load_stats(self):
        try:
            today = jdatetime.date.today()
            today_str = today.strftime("%Y/%m/%d")
            
            week_start = today - jdatetime.timedelta(days=today.weekday())
            week_end = week_start + jdatetime.timedelta(days=6)
            
            month_start = jdatetime.date(today.year, today.month, 1)
            if today.month == 12:
                month_end = jdatetime.date(today.year + 1, 1, 1) - jdatetime.timedelta(days=1)
            else:
                month_end = jdatetime.date(today.year, today.month + 1, 1) - jdatetime.timedelta(days=1)
            
            self.cards["today"].value_label.setText(str(self.session.query(Appointment).filter(Appointment.date == today_str).count()))
            self.cards["week"].value_label.setText(str(self.session.query(Appointment).filter(
                Appointment.date >= week_start.strftime("%Y/%m/%d"),
                Appointment.date <= week_end.strftime("%Y/%m/%d")
            ).count()))
            self.cards["month"].value_label.setText(str(self.session.query(Appointment).filter(
                Appointment.date >= month_start.strftime("%Y/%m/%d"),
                Appointment.date <= month_end.strftime("%Y/%m/%d")
            ).count()))
            self.cards["done"].value_label.setText(str(self.session.query(Appointment).filter(Appointment.status == "انجام شده").count()))
            self.cards["cancelled"].value_label.setText(str(self.session.query(Appointment).filter(Appointment.status.in_(["کنسل استاد", "کنسل مراجع"])).count()))
            self.cards["referred"].value_label.setText(str(self.session.query(Appointment).filter(Appointment.ref_type != "آزاد").count()))
            self.cards["couple"].value_label.setText(str(self.session.query(Appointment).filter(Appointment.patient2_name != "").count()))
            self.cards["doctors"].value_label.setText(str(self.session.query(Doctor).count()))
            self.cards["patients"].value_label.setText(str(self.session.query(Patient).count()))
            self.cards["users"].value_label.setText(str(self.session.query(User).filter(User.is_active == 1).count()))
        except Exception as e:
            print(f"Error: {e}")
