"""
تب مدیریت مالی - نسخه نهایی با دکمه‌های بهینه
"""

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QComboBox,
    QPushButton, QTableWidget, QTableWidgetItem, QTabWidget, QGroupBox,
    QMessageBox, QHeaderView, QAbstractItemView, QFrame,
    QGridLayout, QScrollArea, QFileDialog, QDialog, QMenu
)
from PySide6.QtCore import Qt, QTimer, Signal
from PySide6.QtGui import QAction
from database.engine import SessionLocal
from database.models import ClinicExpense, Appointment, Doctor, DoctorSettlementLog
from sqlalchemy import func
import jdatetime
import os
import tempfile
import webbrowser
import arabic_reshaper
from bidi.algorithm import get_display

from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import matplotlib.pyplot as plt
plt.rcParams['font.family'] = 'Tahoma'

def persian_text(text):
    try:
        reshaped = arabic_reshaper.reshape(str(text))
        return get_display(reshaped)
    except:
        return str(text)

def safe_int(value):
    try:
        if value is None:
            return 0
        return int(float(value))
    except:
        return 0

def export_to_excel(data, headers, title, parent=None):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]
        ws.sheet_view.rightToLeft = True
        
        header_font = Font(name='Tahoma', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1e40af', end_color='1e40af', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        cell_font = Font(name='Tahoma', size=10)
        cell_alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        
        total_font = Font(name='Tahoma', size=11, bold=True, color='FFFFFF')
        total_fill = PatternFill(start_color='2563eb', end_color='2563eb', fill_type='solid')
        
        thin_border = Border(
            left=Side(style='thin', color='cbd5e1'),
            right=Side(style='thin', color='cbd5e1'),
            top=Side(style='thin', color='cbd5e1'),
            bottom=Side(style='thin', color='cbd5e1')
        )
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        totals = {}
        numeric_cols = []
        for col_idx, h in enumerate(headers):
            if any(w in h for w in ["مبلغ", "هزینه", "درآمد", "سهم"]):
                numeric_cols.append(col_idx)
                totals[col_idx] = 0
        
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data):
                cell = ws.cell(row=row_idx, column=col_idx + 1, value=value)
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = thin_border
                
                if col_idx in numeric_cols:
                    try:
                        clean = str(value).replace(',', '').replace('تومان', '').strip()
                        if clean:
                            totals[col_idx] += float(clean)
                    except:
                        pass
        
        for col_idx, header in enumerate(headers, 1):
            max_len = len(header)
            for row_idx in range(2, len(data) + 2):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 35)
        
        if totals:
            total_row = len(data) + 2
            for col_idx in range(len(headers)):
                if col_idx in totals:
                    cell = ws.cell(row=total_row, column=col_idx + 1, value=f"{int(totals[col_idx]):,}")
                elif col_idx == 0:
                    cell = ws.cell(row=total_row, column=1, value="جمع کل")
                else:
                    cell = ws.cell(row=total_row, column=col_idx + 1, value="-")
                
                cell.font = total_font
                cell.fill = total_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
        
        max_row = len(data) + 2 if totals else len(data) + 1
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max_row}"
        ws.freeze_panes = 'A2'
        
        filepath, _ = QFileDialog.getSaveFileName(parent, "ذخیره فایل اکسل", f"{title}.xlsx", "Excel Files (*.xlsx)")
        if filepath:
            wb.save(filepath)
            QMessageBox.information(parent, "موفق", f"✅ فایل ذخیره شد:\n{filepath}")
            return True
    except Exception as e:
        QMessageBox.critical(parent, "خطا", f"خطا: {str(e)}")
    return False

def export_to_pdf(data, headers, title, parent=None):
    try:
        html_lines = []
        html_lines.append('<!DOCTYPE html>')
        html_lines.append('<html dir="rtl">')
        html_lines.append('<head>')
        html_lines.append('<meta charset="UTF-8">')
        html_lines.append(f'<title>{title}</title>')
        html_lines.append('''
        <style>
            @page { size: A4 landscape; margin: 15mm; }
            body { font-family: "Tahoma", "Segoe UI", sans-serif; margin: 0; padding: 20px; background: #f8fafc; }
            .container { background: white; border-radius: 16px; box-shadow: 0 4px 20px rgba(0,0,0,0.1); overflow: hidden; }
            .header { background: linear-gradient(135deg, #1e3a5f, #2563eb); color: white; padding: 20px; text-align: center; }
            .header h1 { margin: 0; font-size: 24px; }
            .header p { margin: 5px 0 0; opacity: 0.9; font-size: 10px; }
            .info { padding: 15px 20px; background: #f1f5f9; display: flex; justify-content: space-between; font-size: 10px; }
            table { width: 100%; border-collapse: collapse; font-size: 10px; }
            th { background: linear-gradient(135deg, #2563eb, #1e40af); color: white; padding: 12px 8px; text-align: center; border: 1px solid #3b82f6; }
            td { padding: 8px; text-align: center; border: 1px solid #cbd5e1; }
            tr:nth-child(even) { background-color: #f8fafc; }
            .total-row { background: #dbeafe; font-weight: bold; }
            .footer { padding: 15px 20px; background: #f1f5f9; display: flex; justify-content: space-between; font-size: 10px; color: #475569; }
            @media print { body { background: white; } .no-print { display: none; } }
        </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>📊 {title}</h1>
                    <p>تاریخ چاپ: ''' + jdatetime.datetime.now().strftime("%Y/%m/%d - %H:%M") + '''</p>
                </div>
                <div class="info">
                    <span>📋 تعداد کل رکوردها: ''' + str(len(data)) + '''</span>
                    <span>🏢 مرکز مشاوره آرامش</span>
                </div>
                <table>
                    <thead>
                        <tr>''')
        
        for h in headers:
            html_lines.append(f'<th>{h}</th>')
        
        html_lines.append('</table></thead><tbody>')
        
        totals = {}
        for row in data:
            html_lines.append('<tr>')
            for col_idx, cell in enumerate(row):
                html_lines.append(f'<td>{cell}</td>')
                if col_idx < len(headers) and any(w in headers[col_idx] for w in ["مبلغ", "سهم", "هزینه"]):
                    try:
                        val = str(cell).replace(',', '').replace('تومان', '').strip()
                        if val:
                            totals[col_idx] = totals.get(col_idx, 0) + float(val)
                    except:
                        pass
            html_lines.append('</tr>')
        
        if totals:
            html_lines.append('<tr class="total-row">')
            for col_idx, h in enumerate(headers):
                if col_idx in totals:
                    html_lines.append(f'<td><strong>جمع: {int(totals[col_idx]):,}</strong></td>')
                elif col_idx == 0:
                    html_lines.append('<td><strong>جمع کل</strong></td>')
                else:
                    html_lines.append('<td>-</td>')
            html_lines.append('</tr>')
        
        html_lines.append('''
                    </tbody>
                </table>
                <div class="footer">
                    <span>🔹 تولید شده توسط سیستم مدیریت مرکز مشاوره آرامش</span>
                    <span>✅ معتبر و رسمی</span>
                </div>
            </div>
            <div class="no-print" style="text-align: center; margin-top: 20px;">
                <button onclick="window.print()" style="background: #3b82f6; color: white; border: none; padding: 10px 20px; border-radius: 8px; cursor: pointer;">🖨️ چاپ / ذخیره PDF</button>
                <button onclick="window.close()" style="background: #ef4444; color: white; border: none; padding: 10px 20px; border-radius: 8px; cursor: pointer; margin-right: 10px;">❌ بستن</button>
            </div>
            <script>setTimeout(function(){ window.print(); }, 500);</script>
        </body>
        </html>
        ''')
        
        html_content = "\n".join(html_lines)
        temp_file = os.path.join(tempfile.gettempdir(), f"{title.replace(' ', '_')}.html")
        with open(temp_file, 'w', encoding='utf-8') as f:
            f.write(html_content)
        webbrowser.open(temp_file)
        return True
    except Exception as e:
        QMessageBox.critical(parent, "خطا", f"خطا: {str(e)}")
        return False

class PersianCalendarDialog(QDialog):
    def __init__(self, parent=None, callback=None):
        super().__init__(parent)
        self.setWindowTitle("انتخاب تاریخ شمسی")
        self.setModal(True)
        self.setFixedSize(550, 520)
        self.callback = callback
        self.current_date = jdatetime.date.today()
        
        layout = QVBoxLayout(self)
        
        header = QLabel("📅 تقویم شمسی")
        header.setStyleSheet("font-size: 18px; font-weight: bold; color: #2563eb; padding: 10px; background: #eff6ff; border-radius: 10px;")
        header.setAlignment(Qt.AlignCenter)
        layout.addWidget(header)
        
        nav_layout = QHBoxLayout()
        self.prev_year_btn = QPushButton("<<")
        self.prev_year_btn.setFixedSize(40, 30)
        self.prev_year_btn.clicked.connect(self.prev_year)
        nav_layout.addWidget(self.prev_year_btn)
        self.prev_month_btn = QPushButton("<")
        self.prev_month_btn.setFixedSize(40, 30)
        self.prev_month_btn.clicked.connect(self.prev_month)
        nav_layout.addWidget(self.prev_month_btn)
        nav_layout.addStretch()
        
        self.year_combo = QComboBox()
        for y in range(1390, jdatetime.date.today().year + 10):
            self.year_combo.addItem(str(y))
        self.year_combo.setCurrentText(str(self.current_date.year))
        self.year_combo.currentTextChanged.connect(self.update_calendar)
        nav_layout.addWidget(self.year_combo)
        
        self.month_combo = QComboBox()
        months = ["فروردین", "اردیبهشت", "خرداد", "تیر", "مرداد", "شهریور", 
                  "مهر", "آبان", "آذر", "دی", "بهمن", "اسفند"]
        self.month_combo.addItems(months)
        self.month_combo.setCurrentIndex(self.current_date.month - 1)
        self.month_combo.currentTextChanged.connect(self.update_calendar)
        nav_layout.addWidget(self.month_combo)
        
        nav_layout.addStretch()
        self.next_month_btn = QPushButton(">")
        self.next_month_btn.setFixedSize(40, 30)
        self.next_month_btn.clicked.connect(self.next_month)
        nav_layout.addWidget(self.next_month_btn)
        self.next_year_btn = QPushButton(">>")
        self.next_year_btn.setFixedSize(40, 30)
        self.next_year_btn.clicked.connect(self.next_year)
        nav_layout.addWidget(self.next_year_btn)
        layout.addLayout(nav_layout)
        
        week_layout = QHBoxLayout()
        week_days = ["شنبه", "یکشنبه", "دوشنبه", "سه‌شنبه", "چهارشنبه", "پنج‌شنبه", "جمعه"]
        for day in week_days:
            lbl = QLabel(day)
            lbl.setAlignment(Qt.AlignCenter)
            lbl.setStyleSheet("font-weight: bold; color: #2563eb; padding: 8px; background: #dbeafe; border-radius: 8px;")
            week_layout.addWidget(lbl)
        layout.addLayout(week_layout)
        
        self.day_grid = QGridLayout()
        self.day_grid.setSpacing(6)
        layout.addLayout(self.day_grid)
        
        today_btn = QPushButton("📅 امروز")
        today_btn.setStyleSheet("background: #10b981; color: white; padding: 8px; border-radius: 10px; font-weight: bold;")
        today_btn.clicked.connect(self.go_today)
        layout.addWidget(today_btn, alignment=Qt.AlignCenter)
        
        self.update_calendar()
    
    def prev_year(self):
        self.current_date = jdatetime.date(self.current_date.year - 1, self.current_date.month, 1)
        self.year_combo.setCurrentText(str(self.current_date.year))
    
    def next_year(self):
        self.current_date = jdatetime.date(self.current_date.year + 1, self.current_date.month, 1)
        self.year_combo.setCurrentText(str(self.current_date.year))
    
    def prev_month(self):
        if self.current_date.month == 1:
            self.current_date = jdatetime.date(self.current_date.year - 1, 12, 1)
        else:
            self.current_date = jdatetime.date(self.current_date.year, self.current_date.month - 1, 1)
        self.year_combo.setCurrentText(str(self.current_date.year))
        self.month_combo.setCurrentIndex(self.current_date.month - 1)
    
    def next_month(self):
        if self.current_date.month == 12:
            self.current_date = jdatetime.date(self.current_date.year + 1, 1, 1)
        else:
            self.current_date = jdatetime.date(self.current_date.year, self.current_date.month + 1, 1)
        self.year_combo.setCurrentText(str(self.current_date.year))
        self.month_combo.setCurrentIndex(self.current_date.month - 1)
    
    def go_today(self):
        self.current_date = jdatetime.date.today()
        self.year_combo.setCurrentText(str(self.current_date.year))
        self.month_combo.setCurrentIndex(self.current_date.month - 1)
    
    def update_calendar(self):
        for i in reversed(range(self.day_grid.count())):
            w = self.day_grid.itemAt(i).widget()
            if w:
                w.deleteLater()
        
        year = int(self.year_combo.currentText())
        month = self.month_combo.currentIndex() + 1
        
        first_day = jdatetime.date(year, month, 1)
        start_weekday = first_day.weekday()
        
        if month <= 6:
            days = 31
        elif month <= 11:
            days = 30
        else:
            days = 30 if first_day.isleap() else 29
        
        row, col = 0, start_weekday
        for day in range(1, days + 1):
            date = jdatetime.date(year, month, day)
            btn = QPushButton(str(day))
            btn.setFixedSize(50, 40)
            if date == jdatetime.date.today():
                btn.setStyleSheet("QPushButton { background: #3b82f6; color: white; border-radius: 10px; font-size: 13px; font-weight: bold; }")
            else:
                btn.setStyleSheet("QPushButton { background: #f1f5f9; color: #1e293b; border-radius: 10px; font-size: 13px; } QPushButton:hover { background: #cbd5e1; }")
            btn.clicked.connect(lambda checked, d=date: self.select_date(d))
            self.day_grid.addWidget(btn, row, col)
            col += 1
            if col > 6:
                col = 0
                row += 1
    
    def select_date(self, date):
        if self.callback:
            self.callback(date.strftime("%Y/%m/%d"))
        self.accept()

class DateSelector(QWidget):
    def __init__(self, label, default_date=None, parent=None):
        super().__init__(parent)
        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(5)
        
        self.label = QLabel(label)
        self.label.setMinimumWidth(40)
        self.label.setStyleSheet("color: #374151; font-size: 10px;")
        layout.addWidget(self.label)
        
        self.date_edit = QLineEdit()
        self.date_edit.setPlaceholderText("1403/01/01")
        self.date_edit.setFixedWidth(130)
        self.date_edit.setFixedHeight(32)
        self.date_edit.setStyleSheet("""
            QLineEdit {
                padding: 5px 8px;
                border: 1px solid #cbd5e1;
                border-radius: 6px;
                background-color: white;
                font-size: 10px;
            }
            QLineEdit:focus {
                border: 2px solid #3b82f6;
            }
        """)
        if default_date:
            self.date_edit.setText(default_date)
        else:
            self.date_edit.setText(jdatetime.date.today().strftime("%Y/%m/%d"))
        layout.addWidget(self.date_edit)
        
        self.calendar_btn = QPushButton("📅")
        self.calendar_btn.setFixedSize(32, 32)
        self.calendar_btn.setCursor(Qt.PointingHandCursor)
        self.calendar_btn.setStyleSheet("""
            QPushButton {
                background-color: #f1f5f9;
                border: 1px solid #cbd5e1;
                border-radius: 6px;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #e2e8f0;
            }
        """)
        self.calendar_btn.clicked.connect(self.show_calendar)
        layout.addWidget(self.calendar_btn)
        
        layout.addStretch()
    
    def show_calendar(self):
        def on_date_selected(date_str):
            self.date_edit.setText(date_str)
        dialog = PersianCalendarDialog(self, on_date_selected)
        dialog.exec()
    
    def get_date(self):
        return self.date_edit.text()
    
    def set_date(self, date_str):
        self.date_edit.setText(date_str)

class FinanceTab(QWidget):
    data_changed = Signal()
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent = parent
        self.session = SessionLocal()
        self.setup_ui()
        self.load_expenses()
        self.load_doctors_list()
        self.update_finance_dashboard()
        self.load_total_stats()
    
    def setup_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(5)
        
        self.inner_tabs = QTabWidget()
        self.inner_tabs.tabBar().setMovable(True)
        
        self.expenses_tab = QWidget()
        self.setup_expenses_tab()
        self.inner_tabs.addTab(self.expenses_tab, "📝 هزینه‌های جاری")
        
        self.finance_report_tab = QWidget()
        self.setup_finance_report_tab()
        self.inner_tabs.addTab(self.finance_report_tab, "📊 گزارش مالی")
        
        self.settlement_tab = QWidget()
        self.setup_settlement_tab()
        self.inner_tabs.addTab(self.settlement_tab, "💰 تسویه اساتید")
        
        self.stats_tab = QWidget()
        self.setup_finance_stats_tab()
        self.inner_tabs.addTab(self.stats_tab, "📈 آمار مالی")
        
        main_layout.addWidget(self.inner_tabs)
    
    def load_total_stats(self):
        try:
            all_apps = self.session.query(Appointment).filter(
                Appointment.is_free == 0,
                Appointment.status != "کنسل استاد",
                Appointment.status != "کنسل مراجع"
            ).all()
            
            total_center_all = 0
            total_doctor_paid_all = 0
            
            for app in all_apps:
                amount = safe_int(app.final_cost)
                try:
                    doc_pct = int(app.doc_share.replace("%", "")) if app.doc_share else 60
                except:
                    doc_pct = 60
                total_center_all += int(amount * (100 - doc_pct) / 100)
                total_doctor_paid_all += int(amount * doc_pct / 100)
            
            self.total_center_all = total_center_all
            self.total_doctor_paid_all = total_doctor_paid_all
            
        except Exception as e:
            print(f"Error: {e}")
            self.total_center_all = 0
            self.total_doctor_paid_all = 0
    
    def setup_expenses_tab(self):
        layout = QVBoxLayout(self.expenses_tab)
        
        form_group = QGroupBox("➕ ثبت هزینه جدید")
        form_group.setStyleSheet("QGroupBox { font-weight: bold; margin-top: 5px; }")
        form_layout = QHBoxLayout(form_group)
        
        form_layout.addWidget(QLabel("تاریخ:"))
        self.exp_date = DateSelector("")
        form_layout.addWidget(self.exp_date)
        
        form_layout.addWidget(QLabel("مبلغ:"))
        self.exp_amount = QLineEdit()
        self.exp_amount.setPlaceholderText("0")
        self.exp_amount.setFixedWidth(100)
        form_layout.addWidget(self.exp_amount)
        
        form_layout.addWidget(QLabel("توضیحات:"))
        self.exp_desc = QLineEdit()
        self.exp_desc.setMinimumWidth(200)
        form_layout.addWidget(self.exp_desc)
        
        self.btn_add_expense = QPushButton("✅ ثبت")
        self.btn_add_expense.setStyleSheet("background: #10b981; color: white; padding: 5px 12px; border-radius: 6px;")
        self.btn_add_expense.clicked.connect(self.add_expense)
        form_layout.addWidget(self.btn_add_expense)
        
        layout.addWidget(form_group)
        
        self.exp_table = QTableWidget()
        self.exp_table.setColumnCount(4)
        self.exp_table.setHorizontalHeaderLabels(["ردیف", "تاریخ", "مبلغ", "توضیحات"])
        self.exp_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.exp_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.exp_table.setAlternatingRowColors(True)
        layout.addWidget(self.exp_table)
        
        btn_layout = QHBoxLayout()
        self.btn_delete_expense = QPushButton("🗑️ حذف")
        self.btn_delete_expense.setStyleSheet("background: #ef4444; color: white; padding: 5px 12px; border-radius: 6px;")
        self.btn_delete_expense.clicked.connect(self.delete_expense)
        btn_layout.addWidget(self.btn_delete_expense)
        
        self.btn_refresh_expenses = QPushButton("🔄 بروزرسانی")
        self.btn_refresh_expenses.setStyleSheet("background: #3b82f6; color: white; padding: 5px 12px; border-radius: 6px;")
        self.btn_refresh_expenses.clicked.connect(self.load_expenses)
        btn_layout.addWidget(self.btn_refresh_expenses)
        btn_layout.addStretch()
        
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("🔍 جستجو:"))
        self.exp_search = QLineEdit()
        self.exp_search.setPlaceholderText("توضیحات...")
        self.exp_search.textChanged.connect(self.search_expenses)
        search_layout.addWidget(self.exp_search)
        search_layout.addStretch()
        
        layout.addLayout(search_layout)
        layout.addLayout(btn_layout)
        
        self.exp_stats = QLabel("")
        self.exp_stats.setStyleSheet("font-size: 10px; color: #6b7280; padding: 3px;")
        layout.addWidget(self.exp_stats)
    
    def load_expenses(self):
        self.exp_table.setRowCount(0)
        expenses = self.session.query(ClinicExpense).order_by(ClinicExpense.date.desc()).all()
        
        total = 0
        for i, exp in enumerate(expenses):
            self.exp_table.insertRow(i)
            amount = safe_int(exp.amount)
            total += amount
            
            self.exp_table.setItem(i, 0, QTableWidgetItem(str(i + 1)))
            self.exp_table.setItem(i, 1, QTableWidgetItem(exp.date or ""))
            self.exp_table.setItem(i, 2, QTableWidgetItem(f"{amount:,}"))
            self.exp_table.setItem(i, 3, QTableWidgetItem(exp.description or ""))
            self.exp_table.item(i, 0).setData(Qt.UserRole, exp.id)
        
        self.exp_stats.setText(f"📊 تعداد: {len(expenses)} | 💰 مجموع: {total:,} تومان")
        self.current_expenses = expenses
    
    def add_expense(self):
        try:
            date_str = self.exp_date.get_date()
            amount_text = self.exp_amount.text().replace(",", "")
            amount = int(amount_text) if amount_text.isdigit() else 0
            
            if amount <= 0:
                QMessageBox.warning(self, "خطا", "مبلغ معتبر وارد کنید")
                return
            
            desc = self.exp_desc.text().strip()
            if not desc:
                QMessageBox.warning(self, "خطا", "توضیحات را وارد کنید")
                return
            
            expense = ClinicExpense(date=date_str, amount=amount, description=desc)
            self.session.add(expense)
            self.session.commit()
            
            self.exp_amount.clear()
            self.exp_desc.clear()
            self.load_expenses()
            self.update_finance_dashboard()
            self.load_total_stats()
            self.data_changed.emit()
            QMessageBox.information(self, "موفق", "هزینه ثبت شد")
            
        except Exception as e:
            self.session.rollback()
            QMessageBox.critical(self, "خطا", f"خطا: {str(e)}")
    
    def delete_expense(self):
        selected = self.exp_table.selectedItems()
        if not selected:
            QMessageBox.warning(self, "خطا", "یک هزینه را انتخاب کنید")
            return
        
        row = selected[0].row()
        exp_id = self.exp_table.item(row, 0).data(Qt.UserRole)
        
        reply = QMessageBox.question(self, "تایید", "آیا از حذف اطمینان دارید؟", QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            try:
                expense = self.session.query(ClinicExpense).get(exp_id)
                if expense:
                    self.session.delete(expense)
                    self.session.commit()
                    self.load_expenses()
                    self.update_finance_dashboard()
                    self.load_total_stats()
                    self.data_changed.emit()
                    QMessageBox.information(self, "موفق", "هزینه حذف شد")
            except Exception as e:
                self.session.rollback()
                QMessageBox.critical(self, "خطا", f"خطا: {str(e)}")
    
    def search_expenses(self, text):
        for row in range(self.exp_table.rowCount()):
            hide = True
            desc_item = self.exp_table.item(row, 3)
            if desc_item and text.lower() in desc_item.text().lower():
                hide = False
            self.exp_table.setRowHidden(row, hide)
    
    def setup_finance_report_tab(self):
        layout = QVBoxLayout(self.finance_report_tab)
        
        filter_frame = QFrame()
        filter_frame.setStyleSheet("QFrame { background: #f8fafc; border-radius: 10px; padding: 8px; }")
        filter_layout = QGridLayout(filter_frame)
        filter_layout.setSpacing(5)
        
        self.report_start = DateSelector("از:", (jdatetime.date.today() - jdatetime.timedelta(days=30)).strftime("%Y/%m/%d"))
        filter_layout.addWidget(self.report_start, 0, 0)
        
        self.report_end = DateSelector("تا:", jdatetime.date.today().strftime("%Y/%m/%d"))
        filter_layout.addWidget(self.report_end, 0, 1)
        
        filter_layout.addWidget(QLabel("استاد:"), 0, 2)
        self.filter_doctor = QComboBox()
        self.filter_doctor.addItem("همه اساتید")
        self.filter_doctor.setMinimumWidth(120)
        filter_layout.addWidget(self.filter_doctor, 0, 3)
        
        btn_layout = QHBoxLayout()
        self.btn_apply_filter = QPushButton("🔍 اعمال فیلتر")
        self.btn_apply_filter.setStyleSheet("background: #3b82f6; color: white; padding: 5px 15px; border-radius: 6px;")
        self.btn_apply_filter.clicked.connect(self.update_finance_dashboard)
        btn_layout.addWidget(self.btn_apply_filter)
        
        self.btn_export_excel = QPushButton("📎 اکسل")
        self.btn_export_excel.setStyleSheet("background: #10b981; color: white; padding: 5px 15px; border-radius: 6px;")
        self.btn_export_excel.clicked.connect(self.export_finance_excel)
        btn_layout.addWidget(self.btn_export_excel)
        
        self.btn_export_pdf = QPushButton("📄 PDF")
        self.btn_export_pdf.setStyleSheet("background: #8b5cf6; color: white; padding: 5px 15px; border-radius: 6px;")
        self.btn_export_pdf.clicked.connect(self.export_finance_pdf)
        btn_layout.addWidget(self.btn_export_pdf)
        
        btn_layout.addStretch()
        filter_layout.addLayout(btn_layout, 1, 0, 1, 4)
        
        layout.addWidget(filter_frame)
        
        summary_frame = QFrame()
        summary_frame.setStyleSheet("QFrame { background: #1e40af; border-radius: 10px; padding: 8px; }")
        summary_layout = QHBoxLayout(summary_frame)
        summary_layout.setSpacing(20)
        
        self.lbl_center_income = QLabel("🏢 سهم مرکز: 0 تومان")
        self.lbl_center_income.setStyleSheet("color: white; font-size: 10px; font-weight: bold;")
        summary_layout.addWidget(self.lbl_center_income)
        
        self.lbl_expenses = QLabel("📝 هزینه‌ها: 0 تومان")
        self.lbl_expenses.setStyleSheet("color: white; font-size: 10px; font-weight: bold;")
        summary_layout.addWidget(self.lbl_expenses)
        
        self.lbl_profit = QLabel("💰 سود خالص: 0 تومان")
        self.lbl_profit.setStyleSheet("color: #bbf7d0; font-size: 10px; font-weight: bold;")
        summary_layout.addWidget(self.lbl_profit)
        
        layout.addWidget(summary_frame)
        
        total_frame = QFrame()
        total_frame.setStyleSheet("QFrame { background: #0f3b5f; border-radius: 10px; padding: 8px; margin-top: 3px; }")
        total_layout = QHBoxLayout(total_frame)
        total_layout.setSpacing(20)
        
        self.lbl_total_center = QLabel("🏢 درآمد کل مرکز از ابتدا: 0 تومان")
        self.lbl_total_center.setStyleSheet("color: #bbf7d0; font-size: 10px; font-weight: bold;")
        total_layout.addWidget(self.lbl_total_center)
        
        self.lbl_total_doctor_paid = QLabel("💰 کل پرداختی به اساتید: 0 تومان")
        self.lbl_total_doctor_paid.setStyleSheet("color: #bbf7d0; font-size: 10px; font-weight: bold;")
        total_layout.addWidget(self.lbl_total_doctor_paid)
        
        layout.addWidget(total_frame)
        
        self.report_table = QTableWidget()
        headers = ["ردیف", "تاریخ", "ساعت", "استاد", "نام مراجع", "مبلغ", "سهم استاد", "سهم مرکز", "وضعیت"]
        self.report_table.setColumnCount(len(headers))
        self.report_table.setHorizontalHeaderLabels(headers)
        self.report_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.report_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.report_table.setAlternatingRowColors(True)
        layout.addWidget(self.report_table)
        
        self.report_stats = QLabel("")
        self.report_stats.setStyleSheet("font-size: 10px; color: #6b7280; padding: 3px;")
        layout.addWidget(self.report_stats)
    
    def load_doctors_list(self):
        doctors = self.session.query(Doctor).order_by(Doctor.name).all()
        for doc in doctors:
            self.filter_doctor.addItem(doc.name)
    
    def update_finance_dashboard(self):
        try:
            start_date = self.report_start.get_date()
            end_date = self.report_end.get_date()
            
            doctor_filter = self.filter_doctor.currentText()
            if doctor_filter == "همه اساتید":
                doctor_filter = None
            
            query = self.session.query(Appointment).filter(
                Appointment.date >= start_date,
                Appointment.date <= end_date,
                Appointment.is_free == 0,
                Appointment.status != "کنسل استاد",
                Appointment.status != "کنسل مراجع"
            )
            if doctor_filter:
                query = query.filter(Appointment.doctor == doctor_filter)
            
            appointments = query.order_by(Appointment.date).all()
            
            total_center_income = 0
            self.report_table.setRowCount(0)
            
            for i, app in enumerate(appointments):
                final_cost = safe_int(app.final_cost)
                try:
                    doc_share_pct = int(app.doc_share.replace("%", "")) if app.doc_share else 60
                except:
                    doc_share_pct = 60
                
                doctor_amount = int(final_cost * doc_share_pct / 100)
                center_amount = int(final_cost * (100 - doc_share_pct) / 100)
                total_center_income += center_amount
                
                self.report_table.insertRow(i)
                self.report_table.setItem(i, 0, QTableWidgetItem(str(i + 1)))
                self.report_table.setItem(i, 1, QTableWidgetItem(app.date or ""))
                self.report_table.setItem(i, 2, QTableWidgetItem(app.time or ""))
                self.report_table.setItem(i, 3, QTableWidgetItem(app.doctor or ""))
                self.report_table.setItem(i, 4, QTableWidgetItem(app.patient_name or ""))
                self.report_table.setItem(i, 5, QTableWidgetItem(f"{final_cost:,}"))
                self.report_table.setItem(i, 6, QTableWidgetItem(f"{doctor_amount:,}"))
                self.report_table.setItem(i, 7, QTableWidgetItem(f"{center_amount:,}"))
                self.report_table.setItem(i, 8, QTableWidgetItem(app.pay_status or "-"))
            
            total_expenses = self.session.query(func.sum(ClinicExpense.amount)).filter(
                ClinicExpense.date >= start_date,
                ClinicExpense.date <= end_date
            ).scalar() or 0
            total_expenses = safe_int(total_expenses)
            net_profit = total_center_income - total_expenses
            
            self.lbl_center_income.setText(f"🏢 سهم مرکز: {total_center_income:,} تومان")
            self.lbl_expenses.setText(f"📝 هزینه‌ها: {total_expenses:,} تومان")
            self.lbl_profit.setText(f"💰 سود خالص: {net_profit:,} تومان")
            
            if hasattr(self, 'total_center_all'):
                self.lbl_total_center.setText(f"🏢 درآمد کل مرکز از ابتدا: {self.total_center_all:,} تومان")
                self.lbl_total_doctor_paid.setText(f"💰 کل پرداختی به اساتید: {self.total_doctor_paid_all:,} تومان")
            
            self.report_stats.setText(f"📊 {len(appointments)} نوبت | 🏢 سهم مرکز: {total_center_income:,}")
            self.current_report_apps = appointments
            
        except Exception as e:
            QMessageBox.critical(self, "خطا", f"خطا: {str(e)}")
    
    def export_finance_excel(self):
        if not hasattr(self, 'current_report_apps') or not self.current_report_apps:
            QMessageBox.warning(self, "خطا", "ابتدا فیلتر را اعمال کنید")
            return
        
        data = []
        for app in self.current_report_apps:
            final_cost = safe_int(app.final_cost)
            try:
                doc_pct = int(app.doc_share.replace("%", "")) if app.doc_share else 60
            except:
                doc_pct = 60
            data.append([
                app.id, app.date, app.time, app.doctor, app.patient_name,
                f"{final_cost:,}", f"{int(final_cost * doc_pct / 100):,}",
                f"{int(final_cost * (100 - doc_pct) / 100):,}", app.pay_status or "-"
            ])
        headers = ["شناسه", "تاریخ", "ساعت", "استاد", "نام مراجع", "مبلغ", "سهم استاد", "سهم مرکز", "وضعیت"]
        export_to_excel(data, headers, "گزارش_مالی", self)
    
    def export_finance_pdf(self):
        if not hasattr(self, 'current_report_apps') or not self.current_report_apps:
            QMessageBox.warning(self, "خطا", "ابتدا فیلتر را اعمال کنید")
            return
        
        data = []
        for app in self.current_report_apps:
            final_cost = safe_int(app.final_cost)
            try:
                doc_pct = int(app.doc_share.replace("%", "")) if app.doc_share else 60
            except:
                doc_pct = 60
            data.append([
                app.id, app.date, app.time, app.doctor, app.patient_name,
                f"{final_cost:,}", f"{int(final_cost * doc_pct / 100):,}",
                f"{int(final_cost * (100 - doc_pct) / 100):,}", app.pay_status or "-"
            ])
        headers = ["شناسه", "تاریخ", "ساعت", "استاد", "نام مراجع", "مبلغ", "سهم استاد", "سهم مرکز", "وضعیت"]
        export_to_pdf(data, headers, "گزارش_مالی", self)
    
    def setup_settlement_tab(self):
        layout = QVBoxLayout(self.settlement_tab)
        
        # ========== بخش تاریخچه تسویه‌ها ==========
        history_group = QGroupBox("📜 تاریخچه تسویه‌ها")
        history_group.setStyleSheet("QGroupBox { font-weight: bold; margin-top: 5px; }")
        history_layout = QVBoxLayout(history_group)
        
        # فیلترهای تاریخچه با دکمه‌های عملیات در کنار هم
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(10)
        
        filter_layout.addWidget(QLabel("از تاریخ:"))
        self.history_start = DateSelector("", (jdatetime.date.today() - jdatetime.timedelta(days=90)).strftime("%Y/%m/%d"))
        filter_layout.addWidget(self.history_start)
        
        filter_layout.addWidget(QLabel("تا تاریخ:"))
        self.history_end = DateSelector("", jdatetime.date.today().strftime("%Y/%m/%d"))
        filter_layout.addWidget(self.history_end)
        
        filter_layout.addWidget(QLabel("استاد:"))
        self.history_doctor = QComboBox()
        self.history_doctor.addItem("همه اساتید")
        filter_layout.addWidget(self.history_doctor)
        
        self.btn_history_filter = QPushButton("🔍 فیلتر")
        self.btn_history_filter.setFixedSize(80, 32)
        self.btn_history_filter.setStyleSheet("background: #3b82f6; color: white; font-weight: bold; border-radius: 6px;")
        self.btn_history_filter.clicked.connect(self.load_settlement_history)
        filter_layout.addWidget(self.btn_history_filter)
        
        # دکمه‌های عملیات تاریخچه (بدون دکمه بروزرسانی/بازنشانی)
        self.btn_history_pdf = QPushButton("📄 گزارش PDF")
        self.btn_history_pdf.setFixedSize(100, 32)
        self.btn_history_pdf.setStyleSheet("background: #8b5cf6; color: white; font-weight: bold; border-radius: 6px;")
        self.btn_history_pdf.clicked.connect(self.on_history_pdf_selected)
        filter_layout.addWidget(self.btn_history_pdf)
        
        self.btn_history_sms = QPushButton("📱 پیامک تایید")
        self.btn_history_sms.setFixedSize(100, 32)
        self.btn_history_sms.setStyleSheet("background: #f59e0b; color: white; font-weight: bold; border-radius: 6px;")
        self.btn_history_sms.clicked.connect(self.on_history_sms_selected)
        filter_layout.addWidget(self.btn_history_sms)
        
        filter_layout.addStretch()
        history_layout.addLayout(filter_layout)
        
        self.history_table = QTableWidget()
        self.history_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.history_table.customContextMenuRequested.connect(self.show_history_context_menu)
        headers_history = ["ردیف", "استاد", "مبلغ", "تاریخ", "تعداد نوبت", "توضیحات"]
        self.history_table.setColumnCount(len(headers_history))
        self.history_table.setHorizontalHeaderLabels(headers_history)
        self.history_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.history_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.history_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.history_table.setAlternatingRowColors(True)
        self.history_table.verticalHeader().setDefaultSectionSize(35)
        history_layout.addWidget(self.history_table)
        
        layout.addWidget(history_group)
        
        # ========== بخش نوبت‌های تسویه نشده ==========
        unsettled_group = QGroupBox("⏳ نوبت‌های تسویه نشده")
        unsettled_group.setStyleSheet("QGroupBox { font-weight: bold; margin-top: 10px; }")
        unsettled_layout = QVBoxLayout(unsettled_group)
        
        # فیلترهای تسویه نشده با دکمه‌های عملیات در کنار هم
        unsettled_filter = QHBoxLayout()
        unsettled_filter.setSpacing(10)
        
        unsettled_filter.addWidget(QLabel("از تاریخ:"))
        self.settle_start = DateSelector("", (jdatetime.date.today() - jdatetime.timedelta(days=30)).strftime("%Y/%m/%d"))
        unsettled_filter.addWidget(self.settle_start)
        
        unsettled_filter.addWidget(QLabel("تا تاریخ:"))
        self.settle_end = DateSelector("", jdatetime.date.today().strftime("%Y/%m/%d"))
        unsettled_filter.addWidget(self.settle_end)
        
        unsettled_filter.addWidget(QLabel("استاد:"))
        self.settle_doctor = QComboBox()
        self.settle_doctor.addItem("همه اساتید")
        unsettled_filter.addWidget(self.settle_doctor)
        
        self.btn_load_settlements = QPushButton("🔄 بروزرسانی")
        self.btn_load_settlements.setFixedSize(80, 32)
        self.btn_load_settlements.setStyleSheet("background: #3b82f6; color: white; font-weight: bold; border-radius: 6px;")
        self.btn_load_settlements.clicked.connect(self.load_settlements)
        unsettled_filter.addWidget(self.btn_load_settlements)
        
        # دکمه تسویه حساب در کنار دکمه بروزرسانی
        self.btn_settle = QPushButton("💰 تسویه حساب")
        self.btn_settle.setFixedSize(100, 32)
        self.btn_settle.setStyleSheet("background: #10b981; color: white; font-weight: bold; border-radius: 6px;")
        self.btn_settle.clicked.connect(self.on_settle_selected)
        unsettled_filter.addWidget(self.btn_settle)
        
        unsettled_filter.addStretch()
        unsettled_layout.addLayout(unsettled_filter)
        
        self.settle_table = QTableWidget()
        self.settle_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.settle_table.customContextMenuRequested.connect(self.show_settle_context_menu)
        headers_settle = ["ردیف", "نام استاد", "تعداد نوبت", "مبلغ قابل پرداخت"]
        self.settle_table.setColumnCount(len(headers_settle))
        self.settle_table.setHorizontalHeaderLabels(headers_settle)
        self.settle_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.settle_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.settle_table.setAlternatingRowColors(True)
        self.settle_table.verticalHeader().setDefaultSectionSize(35)
        unsettled_layout.addWidget(self.settle_table)
        
        unsettled_group.setLayout(unsettled_layout)
        layout.addWidget(unsettled_group)
        
        self.load_doctors_for_history()
        self.load_settlement_history()
    
    def load_doctors_for_history(self):
        doctors = self.session.query(Doctor).order_by(Doctor.name).all()
        for doc in doctors:
            self.history_doctor.addItem(doc.name)
            self.settle_doctor.addItem(doc.name)
    
    def load_settlement_history(self):
        try:
            start_date = self.history_start.get_date()
            end_date = self.history_end.get_date()
            doctor_filter = self.history_doctor.currentText()
            
            query = self.session.query(DoctorSettlementLog)
            
            if start_date and end_date:
                query = query.filter(
                    DoctorSettlementLog.settled_at >= f"{start_date} 00:00:00",
                    DoctorSettlementLog.settled_at <= f"{end_date} 23:59:59"
                )
            
            if doctor_filter != "همه اساتید":
                query = query.filter(DoctorSettlementLog.doctor == doctor_filter)
            
            logs = query.order_by(DoctorSettlementLog.settled_at.desc()).all()
            
            self.history_table.setRowCount(0)
            for i, log in enumerate(logs):
                self.history_table.insertRow(i)
                self.history_table.setItem(i, 0, QTableWidgetItem(str(i + 1)))
                self.history_table.setItem(i, 1, QTableWidgetItem(log.doctor or ""))
                self.history_table.setItem(i, 2, QTableWidgetItem(f"{safe_int(log.amount):,}"))
                self.history_table.setItem(i, 3, QTableWidgetItem(log.settled_at or ""))
                self.history_table.setItem(i, 4, QTableWidgetItem(str(log.appointment_count or 0)))
                self.history_table.setItem(i, 5, QTableWidgetItem(log.description or ""))
            
        except Exception as e:
            print(f"Error: {e}")
    
    def load_settlements(self):
        try:
            start_date = self.settle_start.get_date()
            end_date = self.settle_end.get_date()
            doctor_filter = self.settle_doctor.currentText()
            
            query = self.session.query(Appointment).filter(
                Appointment.date >= start_date,
                Appointment.date <= end_date,
                Appointment.is_free == 0,
                Appointment.status != "کنسل استاد",
                Appointment.status != "کنسل مراجع",
                Appointment.is_settled == 0
            )
            if doctor_filter != "همه اساتید":
                query = query.filter(Appointment.doctor == doctor_filter)
            
            appointments = query.all()
            
            doctor_data = {}
            for app in appointments:
                doc_name = app.doctor
                if doc_name not in doctor_data:
                    doctor_data[doc_name] = {"count": 0, "amount": 0}
                
                final_cost = safe_int(app.final_cost)
                doc_share_text = app.doc_share or "60%"
                try:
                    doc_share_pct = int(doc_share_text.replace("%", ""))
                except:
                    doc_share_pct = 60
                
                doctor_data[doc_name]["count"] += 1
                doctor_data[doc_name]["amount"] += int(final_cost * doc_share_pct / 100)
            
            self.settle_table.setRowCount(0)
            for i, (doc_name, data) in enumerate(doctor_data.items()):
                self.settle_table.insertRow(i)
                self.settle_table.setItem(i, 0, QTableWidgetItem(str(i + 1)))
                self.settle_table.setItem(i, 1, QTableWidgetItem(doc_name))
                self.settle_table.setItem(i, 2, QTableWidgetItem(str(data["count"])))
                self.settle_table.setItem(i, 3, QTableWidgetItem(f"{data['amount']:,}"))
            
            if not doctor_data:
                QMessageBox.information(self, "اطلاع", "هیچ نوبت تسویه نشده‌ای در این بازه وجود ندارد")
                
        except Exception as e:
            QMessageBox.critical(self, "خطا", f"خطا: {str(e)}")
    
    def on_settle_selected(self):
        selected = self.settle_table.selectedItems()
        if not selected:
            QMessageBox.warning(self, "خطا", "لطفاً یک استاد را انتخاب کنید")
            return
        
        row = selected[0].row()
        doctor_name = self.settle_table.item(row, 1).text()
        amount_text = self.settle_table.item(row, 3).text().replace(",", "")
        try:
            amount = int(float(amount_text))
        except:
            amount = 0
        count = int(self.settle_table.item(row, 2).text())
        
        self.settle_selected_doctor(doctor_name, amount, count)
    
    def on_history_pdf_selected(self):
        selected = self.history_table.selectedItems()
        if not selected:
            QMessageBox.warning(self, "خطا", "لطفاً یک رکورد را انتخاب کنید")
            return
        
        row = selected[0].row()
        doctor_name = self.history_table.item(row, 1).text()
        amount_text = self.history_table.item(row, 2).text().replace(",", "")
        try:
            amount = int(float(amount_text))
        except:
            amount = 0
        count = int(self.history_table.item(row, 4).text())
        settled_date = self.history_table.item(row, 3).text()
        
        self.export_settlement_pdf(doctor_name, amount, count, settled_date)
    
    def on_history_sms_selected(self):
        selected = self.history_table.selectedItems()
        if not selected:
            QMessageBox.warning(self, "خطا", "لطفاً یک رکورد را انتخاب کنید")
            return
        
        row = selected[0].row()
        doctor_name = self.history_table.item(row, 1).text()
        amount_text = self.history_table.item(row, 2).text().replace(",", "")
        try:
            amount = int(float(amount_text))
        except:
            amount = 0
        count = int(self.history_table.item(row, 4).text())
        
        self.send_sms_to_doctor_from_history(doctor_name, amount, count)
    
    def send_sms_to_doctor_from_history(self, doctor_name, amount, count):
        logs = self.session.query(DoctorSettlementLog).filter(
            DoctorSettlementLog.doctor == doctor_name,
            DoctorSettlementLog.amount == amount,
            DoctorSettlementLog.appointment_count == count
        ).first()
        
        if logs:
            start_date = logs.start_date
            end_date = logs.end_date
            apps = self.session.query(Appointment).filter(
                Appointment.date >= start_date,
                Appointment.date <= end_date,
                Appointment.doctor == doctor_name,
                Appointment.is_free == 0
            ).all()
        else:
            apps = []
        
        message = f"""بسم الله الرحمن الرحیم

استاد گرامی {doctor_name}

سلام علیکم

احتراماً، به اطلاع می‌رساند گزارش تسویه حساب شما به شرح زیر می‌باشد:

📋 خلاصه تسویه:
━━━━━━━━━━━━━━━━━━━━
🔹 تعداد جلسات: {count} نوبت
💰 مبلغ کل تسویه شده: {amount:,} تومان
📅 تاریخ تسویه: {jdatetime.datetime.now().strftime("%Y/%m/%d")}

📊 جزئیات جلسات:
━━━━━━━━━━━━━━━━━━━━
"""
        
        for i, app in enumerate(apps, 1):
            subject = app.subject or "مشاوره"
            message += f"""
{i}. تاریخ: {app.date} - ساعت: {app.time}
   مراجع: {app.patient_name}
   موضوع: {subject}
   مبلغ جلسه: {safe_int(app.final_cost):,} تومان
   سهم استاد ({app.doc_share or '60%'}): {int(safe_int(app.final_cost) * int((app.doc_share or '60').replace('%', '')) / 100):,} تومان
━━━━━━━━━━━━━━━━━━━━
"""
        
        message += f"""
✅ جمع کل مبلغ تسویه شده: {amount:,} تومان

🙏 بدین وسیله تأیید می‌گردد که مبلغ فوق بابت تسویه حساب جلسات شما به حساب شما واریز گردید.

با تشکر و احترام
مدیریت مرکز مشاوره آرامش
━━━━━━━━━━━━━━━━━━━━
🔹 این پیام به صورت خودکار ارسال شده است"""
        
        from PySide6.QtGui import QGuiApplication
        clipboard = QGuiApplication.clipboard()
        clipboard.setText(message)
        
        QMessageBox.information(self, "پیامک", 
            "✅ متن پیامک تایید تسویه در حافظه کپی شد.\n\n"
            "اکنون می‌توانید آن را در پیام رسان مورد نظر paste کنید و برای استاد ارسال نمایید.")
    
    def export_settlement_pdf(self, doctor_name, amount, count, settled_date):
        logs = self.session.query(DoctorSettlementLog).filter(
            DoctorSettlementLog.doctor == doctor_name,
            DoctorSettlementLog.amount == amount,
            DoctorSettlementLog.appointment_count == count
        ).first()
        
        if logs:
            start_date = logs.start_date
            end_date = logs.end_date
            apps = self.session.query(Appointment).filter(
                Appointment.date >= start_date,
                Appointment.date <= end_date,
                Appointment.doctor == doctor_name,
                Appointment.is_free == 0
            ).all()
        else:
            apps = []
        
        html_lines = []
        html_lines.append('<!DOCTYPE html>')
        html_lines.append('<html dir="rtl">')
        html_lines.append('<head>')
        html_lines.append('<meta charset="UTF-8">')
        html_lines.append(f'<title>گزارش تسویه حساب استاد {doctor_name}</title>')
        html_lines.append('''
        <style>
            @page { size: A4; margin: 20mm; }
            body { font-family: "Tahoma", "Segoe UI", sans-serif; margin: 0; padding: 20px; background: #f8fafc; }
            .container { background: white; border-radius: 16px; box-shadow: 0 4px 20px rgba(0,0,0,0.1); overflow: hidden; }
            .header { background: linear-gradient(135deg, #1e3a5f, #2563eb); color: white; padding: 25px; text-align: center; }
            .header h1 { margin: 0; font-size: 24px; }
            .header p { margin: 5px 0 0; opacity: 0.9; }
            .info { padding: 15px 20px; background: #f1f5f9; display: flex; justify-content: space-between; }
            table { width: 100%; border-collapse: collapse; margin: 15px 0; }
            th { background: #2563eb; color: white; padding: 12px; border: 1px solid #3b82f6; }
            td { padding: 10px; border: 1px solid #cbd5e1; text-align: center; }
            tr:nth-child(even) { background: #f8fafc; }
            .total { background: #dbeafe; font-weight: bold; padding: 15px; text-align: center; }
            .footer { padding: 15px; background: #f1f5f9; text-align: center; font-size: 10px; }
            .greeting { padding: 20px; line-height: 2; }
        </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>📄 گزارش تسویه حساب</h1>
                    <p>مرکز مشاوره آرامش</p>
                </div>
                <div class="greeting">
                    <p>استاد گرامی ''' + doctor_name + '''</p>
                    <p>سلام علیکم</p>
                    <p>احتراماً، گزارش تسویه حساب شما به شرح زیر می‌باشد:</p>
                </div>
                <div class="info">
                    <span>📅 تاریخ تسویه: ''' + settled_date + '''</span>
                    <span>💰 مبلغ کل تسویه شده: ''' + f"{amount:,}" + ''' تومان</span>
                    <span>📋 تعداد جلسات: ''' + str(count) + ''' نوبت</span>
                </div>
                <table>
                    <thead>
                        <tr>
                            <th>ردیف</th>
                            <th>تاریخ</th>
                            <th>ساعت</th>
                            <th>نام مراجع</th>
                            <th>موضوع</th>
                            <th>مبلغ جلسه</th>
                            <th>سهم استاد</th>
                        </tr>
                    </thead>
                    <tbody>
        ''')
        
        for i, app in enumerate(apps, 1):
            subject = app.subject or "مشاوره"
            amount_app = safe_int(app.final_cost)
            doc_share_pct = int((app.doc_share or '60').replace('%', ''))
            doctor_share = int(amount_app * doc_share_pct / 100)
            html_lines.append(f"""
                        <tr>
                            <td>{i}拥有
                            <td>{app.date}拥有
                            <td>{app.time}拥有
                            <td>{app.patient_name}拥有
                            <td>{subject}拥有
                            <td>{amount_app:,} تومان拥有
                            <td>{doctor_share:,} تومان ({doc_share_pct}%)拥有
                        </tr>
            """)
        
        html_lines.append(f"""
                    </tbody>
                </table>
                <div class="total">
                    <strong>جمع کل مبلغ تسویه شده: {amount:,} تومان</strong>
                </div>
                <div class="greeting">
                    <p>✅ بدین وسیله تأیید می‌گردد که مبلغ فوق بابت تسویه حساب جلسات شما به حساب شما واریز گردید.</p>
                    <p>با تشکر و احترام<br>مدیریت مرکز مشاوره آرامش</p>
                </div>
                <div class="footer">
                    <p>🔹 این گزارش توسط سیستم مدیریت مرکز مشاوره آرامش تولید شده است</p>
                    <p>تاریخ چاپ: {jdatetime.datetime.now().strftime("%Y/%m/%d - %H:%M")}</p>
                </div>
            </div>
        </body>
        </html>
        """)
        
        html_content = "\n".join(html_lines)
        temp_file = os.path.join(tempfile.gettempdir(), f"settlement_{doctor_name}.html")
        with open(temp_file, 'w', encoding='utf-8') as f:
            f.write(html_content)
        webbrowser.open(temp_file)
        
        QMessageBox.information(self, "PDF", 
            "✅ فایل PDF گزارش تسویه در مرورگر باز شد.\n\n"
            "می‌توانید آن را ذخیره کرده و برای استاد ارسال نمایید.")
    
    def send_sms_to_doctor_from_unsettled(self, doctor_name, amount, count):
        start_date = self.settle_start.get_date()
        end_date = self.settle_end.get_date()
        
        apps = self.session.query(Appointment).filter(
            Appointment.date >= start_date,
            Appointment.date <= end_date,
            Appointment.doctor == doctor_name,
            Appointment.is_free == 0,
            Appointment.is_settled == 0
        ).all()
        
        message = f"""بسم الله الرحمن الرحیم

استاد گرامی {doctor_name}

سلام علیکم

احتراماً، به اطلاع می‌رساند صورتجلسات تسویه نشده شما به شرح زیر می‌باشد:

📋 خلاصه:
━━━━━━━━━━━━━━━━━━━━
🔹 تعداد جلسات تسویه نشده: {count} نوبت
💰 مبلغ کل قابل پرداخت: {amount:,} تومان
📅 بازه زمانی: {start_date} تا {end_date}

📊 جزئیات جلسات:
━━━━━━━━━━━━━━━━━━━━
"""
        
        for i, app in enumerate(apps, 1):
            subject = app.subject or "مشاوره"
            message += f"""
{i}. تاریخ: {app.date} - ساعت: {app.time}
   مراجع: {app.patient_name}
   موضوع: {subject}
   مبلغ جلسه: {safe_int(app.final_cost):,} تومان
   سهم استاد ({app.doc_share or '60%'}): {int(safe_int(app.final_cost) * int((app.doc_share or '60').replace('%', '')) / 100):,} تومان
━━━━━━━━━━━━━━━━━━━━
"""
        
        message += f"""
✅ جمع کل مبلغ قابل پرداخت: {amount:,} تومان

🙏 لطفاً جهت تسویه حساب، مبلغ فوق را به شماره کارت یا شماره شبا اعلام شده واریز فرمایید.

پس از واریز، تصویر فیش واریزی را در پنل کاربری آپلود کنید.

با تشکر و احترام
مدیریت مرکز مشاوره آرامش
━━━━━━━━━━━━━━━━━━━━
🔹 این پیام به صورت خودکار ارسال شده است"""
        
        from PySide6.QtGui import QGuiApplication
        clipboard = QGuiApplication.clipboard()
        clipboard.setText(message)
        
        QMessageBox.information(self, "پیامک", 
            "✅ متن پیامک در حافظه کپی شد.\n\n"
            "اکنون می‌توانید آن را در پیام رسان مورد نظر paste کنید و برای استاد ارسال نمایید.")
    
    def settle_selected_doctor(self, doctor_name, amount, count):
        reply = QMessageBox.question(self, "تایید تسویه",
            f"آیا از تسویه حساب استاد «{doctor_name}» به مبلغ {amount:,} تومان برای {count} نوبت اطمینان دارید؟",
            QMessageBox.Yes | QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            try:
                start_date = self.settle_start.get_date()
                end_date = self.settle_end.get_date()
                
                apps = self.session.query(Appointment).filter(
                    Appointment.date >= start_date,
                    Appointment.date <= end_date,
                    Appointment.doctor == doctor_name,
                    Appointment.is_free == 0,
                    Appointment.status != "کنسل استاد",
                    Appointment.status != "کنسل مراجع",
                    Appointment.is_settled == 0
                ).all()
                
                for app in apps:
                    app.is_settled = 1
                
                log = DoctorSettlementLog(
                    doctor=doctor_name,
                    amount=amount,
                    start_date=start_date,
                    end_date=end_date,
                    settled_at=jdatetime.datetime.now().strftime("%Y/%m/%d %H:%M:%S"),
                    appointment_count=count,
                    description=f"تسویه حساب استاد {doctor_name}"
                )
                self.session.add(log)
                self.session.commit()
                
                QMessageBox.information(self, "موفق", f"تسویه حساب استاد {doctor_name} با موفقیت انجام شد")
                self.load_settlements()
                self.load_settlement_history()
                self.update_finance_dashboard()
                self.load_total_stats()
                self.data_changed.emit()
                
            except Exception as e:
                self.session.rollback()
                QMessageBox.critical(self, "خطا", f"خطا در تسویه حساب:\n{str(e)}")
    
    def setup_finance_stats_tab(self):
        layout = QVBoxLayout(self.stats_tab)
        
        filter_frame = QFrame()
        filter_frame.setStyleSheet("QFrame { background: #f8fafc; border-radius: 10px; padding: 8px; }")
        filter_layout = QHBoxLayout(filter_frame)
        
        filter_layout.addWidget(QLabel("سال:"))
        self.stats_year = QComboBox()
        for y in range(1400, jdatetime.date.today().year + 2):
            self.stats_year.addItem(str(y))
        self.stats_year.setCurrentText(str(jdatetime.date.today().year))
        filter_layout.addWidget(self.stats_year)
        
        self.stats_type = QComboBox()
        self.stats_type.addItems(["درآمد ماهیانه", "سهم مرکز ماهیانه", "تعداد نوبت‌ها ماهیانه"])
        filter_layout.addWidget(self.stats_type)
        
        self.stats_btn = QPushButton("📊 رسم نمودار")
        self.stats_btn.setStyleSheet("background: #3b82f6; color: white; padding: 5px 15px; border-radius: 6px;")
        self.stats_btn.clicked.connect(self.draw_finance_chart)
        filter_layout.addWidget(self.stats_btn)
        
        filter_layout.addStretch()
        layout.addWidget(filter_frame)
        
        self.chart_frame = QFrame()
        self.chart_frame.setStyleSheet("QFrame { background: white; border-radius: 10px; }")
        self.chart_layout = QVBoxLayout(self.chart_frame)
        self.chart_label = QLabel("📈 روی دکمه رسم نمودار کلیک کنید")
        self.chart_label.setAlignment(Qt.AlignCenter)
        self.chart_label.setStyleSheet("color: #6b7280; padding: 50px;")
        self.chart_layout.addWidget(self.chart_label)
        
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setWidget(self.chart_frame)
        layout.addWidget(scroll)
        
        self.current_fig = None
    
    def draw_finance_chart(self):
        for i in reversed(range(self.chart_layout.count())):
            w = self.chart_layout.itemAt(i).widget()
            if w:
                w.deleteLater()
        
        year = int(self.stats_year.currentText())
        chart_type = self.stats_type.currentText()
        
        try:
            fig = Figure(figsize=(10, 5), facecolor='white')
            ax = fig.add_subplot(111)
            
            month_names = ["فروردین", "اردیبهشت", "خرداد", "تیر", "مرداد", "شهریور", 
                          "مهر", "آبان", "آذر", "دی", "بهمن", "اسفند"]
            persian_months = [persian_text(m) for m in month_names]
            values = []
            
            for month in range(1, 13):
                month_start = jdatetime.date(year, month, 1)
                if month == 12:
                    month_end = jdatetime.date(year + 1, 1, 1) - jdatetime.timedelta(days=1)
                else:
                    month_end = jdatetime.date(year, month + 1, 1) - jdatetime.timedelta(days=1)
                
                if chart_type == "درآمد ماهیانه":
                    apps = self.session.query(Appointment).filter(
                        Appointment.date >= month_start.strftime("%Y/%m/%d"),
                        Appointment.date <= month_end.strftime("%Y/%m/%d"),
                        Appointment.is_free == 0                    ).all()
                    total = sum(safe_int(a.final_cost) for a in apps)
                    values.append(total)
                    
                elif chart_type == "سهم مرکز ماهیانه":
                    apps = self.session.query(Appointment).filter(
                        Appointment.date >= month_start.strftime("%Y/%m/%d"),
                        Appointment.date <= month_end.strftime("%Y/%m/%d"),
                        Appointment.is_free == 0
                    ).all()
                    total = 0
                    for app in apps:
                        amount = safe_int(app.final_cost)
                        try:
                            pct = int(app.doc_share.replace("%", "")) if app.doc_share else 60
                        except:
                            pct = 60
                        total += int(amount * (100 - pct) / 100)
                    values.append(total)
                    
                else:
                    apps = self.session.query(Appointment).filter(
                        Appointment.date >= month_start.strftime("%Y/%m/%d"),
                        Appointment.date <= month_end.strftime("%Y/%m/%d")
                    ).count()
                    values.append(apps)
            
            ax.bar(persian_months, values, color='#3b82f6')
            ax.set_xlabel(persian_text('ماه'))
            if "درآمد" in chart_type or "سهم" in chart_type:
                ax.set_ylabel(persian_text('مبلغ (تومان)'))
            else:
                ax.set_ylabel(persian_text('تعداد'))
            ax.set_title(persian_text(f'{chart_type} - سال {year}'))
            plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha='right')
            
            fig.tight_layout()
            canvas = FigureCanvas(fig)
            self.current_fig = fig
            self.chart_layout.addWidget(canvas)
            
        except Exception as e:
            self.chart_layout.addWidget(self.chart_label)
            QMessageBox.critical(self, "خطا", f"خطا در رسم نمودار:\n{str(e)}")
    
    def show_settle_context_menu(self, position):
        item = self.settle_table.itemAt(position)
        if not item:
            return
        
        row = item.row()
        doctor_name = self.settle_table.item(row, 1).text()
        amount_text = self.settle_table.item(row, 3).text().replace(",", "")
        try:
            amount = int(float(amount_text))
        except:
            amount = 0
        count = int(self.settle_table.item(row, 2).text())
        
        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu {
                background-color: white;
                border: 1px solid #cbd5e1;
                border-radius: 8px;
                padding: 5px;
            }
            QMenu::item {
                padding: 8px 25px;
                margin: 2px;
                border-radius: 4px;
            }
            QMenu::item:selected {
                background-color: #3b82f6;
                color: white;
            }
        """)
        
        action_settle = QAction("💰 تسویه حساب", self)
        action_settle.triggered.connect(lambda: self.settle_selected_doctor(doctor_name, amount, count))
        menu.addAction(action_settle)
        
        menu.exec(self.settle_table.viewport().mapToGlobal(position))
    
    def show_history_context_menu(self, position):
        item = self.history_table.itemAt(position)
        if not item:
            return
        
        row = item.row()
        doctor_name = self.history_table.item(row, 1).text()
        amount_text = self.history_table.item(row, 2).text().replace(",", "")
        try:
            amount = int(float(amount_text))
        except:
            amount = 0
        count = int(self.history_table.item(row, 4).text())
        settled_date = self.history_table.item(row, 3).text()
        
        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu {
                background-color: white;
                border: 1px solid #cbd5e1;
                border-radius: 8px;
                padding: 5px;
            }
            QMenu::item {
                padding: 8px 25px;
                margin: 2px;
                border-radius: 4px;
            }
            QMenu::item:selected {
                background-color: #3b82f6;
                color: white;
            }
        """)
        
        action_pdf = QAction("📄 گزارش PDF", self)
        action_pdf.triggered.connect(lambda: self.export_settlement_pdf(doctor_name, amount, count, settled_date))
        menu.addAction(action_pdf)
        
        action_sms = QAction("📱 پیامک تایید", self)
        action_sms.triggered.connect(lambda: self.send_sms_to_doctor_from_history(doctor_name, amount, count))
        menu.addAction(action_sms)
        
        menu.exec(self.history_table.viewport().mapToGlobal(position))

    def __del__(self):
        if hasattr(self, 'session'):
            try:
                self.session.close()
            except:
                pass