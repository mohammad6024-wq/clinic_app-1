"""
تب گزارشات حرفه‌ای - نسخه نهایی با خروجی‌های حرفه‌ای
"""

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QComboBox, QPushButton,
    QFileDialog, QMessageBox, QFrame, QGridLayout,
    QLineEdit, QTableWidget, QTableWidgetItem, QHeaderView,
    QTabWidget, QCompleter, QScrollArea, QDialog
)
from PySide6.QtCore import Qt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from database.engine import SessionLocal
from database.models import Appointment, Doctor, Patient, Subject
import jdatetime
import matplotlib.pyplot as plt
import os
import tempfile
import webbrowser
import arabic_reshaper
from bidi.algorithm import get_display

plt.rcParams['font.family'] = 'Tahoma'

def persian_text(text):
    if not text:
        return ""
    try:
        reshaped = arabic_reshaper.reshape(str(text))
        return get_display(reshaped)
    except:
        return str(text)

def safe_int(value):
    try:
        if value is None:
            return 0
        return int(float(value))
    except:
        return 0

# ==================== خروجی اکسل حرفه‌ای ====================
def export_to_excel(data, headers, title, parent=None):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]
        ws.sheet_view.rightToLeft = True
        
        # استایل هدر
        header_font = Font(name='Tahoma', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1e40af', end_color='1e40af', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # استایل سلول
        cell_font = Font(name='Tahoma', size=10)
        cell_alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        
        # استایل جمع
        total_font = Font(name='Tahoma', size=11, bold=True, color='FFFFFF')
        total_fill = PatternFill(start_color='2563eb', end_color='2563eb', fill_type='solid')
        
        thin_border = Border(
            left=Side(style='thin', color='cbd5e1'),
            right=Side(style='thin', color='cbd5e1'),
            top=Side(style='thin', color='cbd5e1'),
            bottom=Side(style='thin', color='cbd5e1')
        )
        
        # هدرها
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # داده‌ها
        totals = {}
        numeric_cols = []
        for col_idx, h in enumerate(headers):
            if any(w in h for w in ["مبلغ", "سهم", "هزینه", "درآمد"]):
                numeric_cols.append(col_idx)
                totals[col_idx] = 0
        
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data):
                cell = ws.cell(row=row_idx, column=col_idx + 1, value=value)
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = thin_border
                
                if col_idx in numeric_cols:
                    try:
                        clean = str(value).replace(',', '').replace('تومان', '').strip()
                        if clean:
                            totals[col_idx] += float(clean)
                    except:
                        pass
        
        # عرض ستون‌ها
        for col_idx, header in enumerate(headers, 1):
            max_len = len(header)
            for row_idx in range(2, len(data) + 2):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 35)
        
        # ردیف جمع
        if totals:
            total_row = len(data) + 2
            for col_idx in range(len(headers)):
                if col_idx in totals:
                    cell = ws.cell(row=total_row, column=col_idx + 1, value=f"{int(totals[col_idx]):,}")
                elif col_idx == 0:
                    cell = ws.cell(row=total_row, column=1, value="جمع کل")
                else:
                    cell = ws.cell(row=total_row, column=col_idx + 1, value="-")
                
                cell.font = total_font
                cell.fill = total_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
        
        # AutoFilter
        max_row = len(data) + 2 if totals else len(data) + 1
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max_row}"
        ws.freeze_panes = 'A2'
        
        filepath, _ = QFileDialog.getSaveFileName(parent, "ذخیره فایل اکسل", f"{title}.xlsx", "Excel Files (*.xlsx)")
        if filepath:
            wb.save(filepath)
            QMessageBox.information(parent, "موفق", f"✅ فایل ذخیره شد:\n{filepath}")
            return True
    except Exception as e:
        QMessageBox.critical(parent, "خطا", f"خطا: {str(e)}")
    return False

# ==================== خروجی PDF حرفه‌ای ====================
def export_to_pdf(data, headers, title, parent=None):
    try:
        html = f"""
        <!DOCTYPE html>
        <html dir="rtl">
        <head>
            <meta charset="utf-8">
            <title>{title}</title>
            <style>
                @page {{ size: A4 landscape; margin: 15mm; }}
                body {{ font-family: 'Tahoma', sans-serif; margin: 0; padding: 20px; background: #f8fafc; }}
                .container {{ background: white; border-radius: 16px; box-shadow: 0 4px 20px rgba(0,0,0,0.1); overflow: hidden; }}
                .header {{ background: linear-gradient(135deg, #1e3a5f, #2563eb); color: white; padding: 20px; text-align: center; }}
                .header h1 {{ margin: 0; font-size: 24px; }}
                .header p {{ margin: 5px 0 0; opacity: 0.9; font-size: 12px; }}
                .info {{ padding: 15px 20px; background: #f1f5f9; display: flex; justify-content: space-between; font-size: 12px; }}
                table {{ width: 100%; border-collapse: collapse; font-size: 11px; }}
                th {{ background: linear-gradient(135deg, #2563eb, #1e40af); color: white; padding: 12px 8px; text-align: center; border: 1px solid #3b82f6; }}
                td {{ padding: 8px; text-align: center; border: 1px solid #cbd5e1; }}
                tr:nth-child(even) {{ background-color: #f8fafc; }}
                .total-row {{ background: #dbeafe; font-weight: bold; }}
                .footer {{ padding: 15px 20px; background: #f1f5f9; display: flex; justify-content: space-between; font-size: 11px; color: #475569; }}
                @media print {{ body {{ background: white; }} .no-print {{ display: none; }} }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>📊 {title}</h1>
                    <p>تاریخ چاپ: {jdatetime.datetime.now().strftime("%Y/%m/%d - %H:%M")}</p>
                </div>
                <div class="info">
                    <span>📋 تعداد کل رکوردها: {len(data)}</span>
                    <span>🏢 مرکز مشاوره آرامش</span>
                </div>
                <table>
                    <thead><tr>"""
        for h in headers:
            html += f"<th>{h}</th>"
        html += "</tr></thead><tbody>"
        
        totals = {}
        for row in data:
            html += "<tr>"
            for col_idx, cell in enumerate(row):
                html += f"<td>{cell}</td>"
                if col_idx < len(headers) and any(w in headers[col_idx] for w in ["مبلغ", "سهم", "هزینه"]):
                    try:
                        val = str(cell).replace(',', '').replace('تومان', '').strip()
                        if val:
                            totals[col_idx] = totals.get(col_idx, 0) + float(val)
                    except:
                        pass
            html += "</tr>"
        
        if totals:
            html += '<tr class="total-row">'
            for col_idx, h in enumerate(headers):
                if col_idx in totals:
                    html += f"<td><strong>جمع: {int(totals[col_idx]):,}</strong></td>"
                elif col_idx == 0:
                    html += "<td><strong>جمع کل</strong></td>"
                else:
                    html += "<td>-</td>"
            html += "</tr>"
        
        html += """
                    </tbody>
                </table>
                <div class="footer">
                    <span>🔹 تولید شده توسط سیستم مدیریت مرکز مشاوره آرامش</span>
                    <span>✅ معتبر و رسمی</span>
                </div>
            </div>
            <div class="no-print" style="text-align: center; margin-top: 20px;">
                <button onclick="window.print()" style="background: #3b82f6; color: white; border: none; padding: 10px 20px; border-radius: 8px; cursor: pointer;">🖨️ چاپ / ذخیره PDF</button>
                <button onclick="window.close()" style="background: #ef4444; color: white; border: none; padding: 10px 20px; border-radius: 8px; cursor: pointer; margin-right: 10px;">❌ بستن</button>
            </div>
            <script>setTimeout(function(){{ window.print(); }}, 500);</script>
        </body>
        </html>
        """
        
        temp_file = os.path.join(tempfile.gettempdir(), f"{title.replace(' ', '_')}.html")
        with open(temp_file, 'w', encoding='utf-8') as f:
            f.write(html)
        webbrowser.open(temp_file)
        return True
    except Exception as e:
        QMessageBox.critical(parent, "خطا", f"خطا: {str(e)}")
        return False

# ==================== تقویم شمسی ====================
class PersianCalendarDialog(QDialog):
    def __init__(self, parent=None, callback=None):
        super().__init__(parent)
        self.setWindowTitle("انتخاب تاریخ شمسی")
        self.setModal(True)
        self.setFixedSize(550, 520)
        self.callback = callback
        self.current_date = jdatetime.date.today()
        
        layout = QVBoxLayout(self)
        
        header = QLabel("📅 تقویم شمسی")
        header.setStyleSheet("font-size: 18px; font-weight: bold; color: #2563eb; padding: 10px; background: #eff6ff; border-radius: 10px;")
        header.setAlignment(Qt.AlignCenter)
        layout.addWidget(header)
        
        nav_layout = QHBoxLayout()
        self.prev_year_btn = QPushButton("<<")
        self.prev_year_btn.setFixedSize(40, 30)
        self.prev_year_btn.clicked.connect(self.prev_year)
        nav_layout.addWidget(self.prev_year_btn)
        self.prev_month_btn = QPushButton("<")
        self.prev_month_btn.setFixedSize(40, 30)
        self.prev_month_btn.clicked.connect(self.prev_month)
        nav_layout.addWidget(self.prev_month_btn)
        nav_layout.addStretch()
        
        self.year_combo = QComboBox()
        for y in range(1390, jdatetime.date.today().year + 10):
            self.year_combo.addItem(str(y))
        self.year_combo.setCurrentText(str(self.current_date.year))
        self.year_combo.currentTextChanged.connect(self.update_calendar)
        nav_layout.addWidget(self.year_combo)
        
        self.month_combo = QComboBox()
        months = ["فروردین", "اردیبهشت", "خرداد", "تیر", "مرداد", "شهریور", 
                  "مهر", "آبان", "آذر", "دی", "بهمن", "اسفند"]
        self.month_combo.addItems(months)
        self.month_combo.setCurrentIndex(self.current_date.month - 1)
        self.month_combo.currentTextChanged.connect(self.update_calendar)
        nav_layout.addWidget(self.month_combo)
        
        nav_layout.addStretch()
        self.next_month_btn = QPushButton(">")
        self.next_month_btn.setFixedSize(40, 30)
        self.next_month_btn.clicked.connect(self.next_month)
        nav_layout.addWidget(self.next_month_btn)
        self.next_year_btn = QPushButton(">>")
        self.next_year_btn.setFixedSize(40, 30)
        self.next_year_btn.clicked.connect(self.next_year)
        nav_layout.addWidget(self.next_year_btn)
        layout.addLayout(nav_layout)
        
        week_layout = QHBoxLayout()
        week_days = ["شنبه", "یکشنبه", "دوشنبه", "سه‌شنبه", "چهارشنبه", "پنج‌شنبه", "جمعه"]
        for day in week_days:
            lbl = QLabel(day)
            lbl.setAlignment(Qt.AlignCenter)
            lbl.setStyleSheet("font-weight: bold; color: #2563eb; padding: 8px; background: #dbeafe; border-radius: 8px;")
            week_layout.addWidget(lbl)
        layout.addLayout(week_layout)
        
        self.day_grid = QGridLayout()
        self.day_grid.setSpacing(6)
        layout.addLayout(self.day_grid)
        
        today_btn = QPushButton("📅 امروز")
        today_btn.setStyleSheet("background: #10b981; color: white; padding: 8px; border-radius: 10px; font-weight: bold;")
        today_btn.clicked.connect(self.go_today)
        layout.addWidget(today_btn, alignment=Qt.AlignCenter)
        
        self.update_calendar()
    
    def prev_year(self):
        self.current_date = jdatetime.date(self.current_date.year - 1, self.current_date.month, 1)
        self.year_combo.setCurrentText(str(self.current_date.year))
    
    def next_year(self):
        self.current_date = jdatetime.date(self.current_date.year + 1, self.current_date.month, 1)
        self.year_combo.setCurrentText(str(self.current_date.year))
    
    def prev_month(self):
        if self.current_date.month == 1:
            self.current_date = jdatetime.date(self.current_date.year - 1, 12, 1)
        else:
            self.current_date = jdatetime.date(self.current_date.year, self.current_date.month - 1, 1)
        self.year_combo.setCurrentText(str(self.current_date.year))
        self.month_combo.setCurrentIndex(self.current_date.month - 1)
    
    def next_month(self):
        if self.current_date.month == 12:
            self.current_date = jdatetime.date(self.current_date.year + 1, 1, 1)
        else:
            self.current_date = jdatetime.date(self.current_date.year, self.current_date.month + 1, 1)
        self.year_combo.setCurrentText(str(self.current_date.year))
        self.month_combo.setCurrentIndex(self.current_date.month - 1)
    
    def go_today(self):
        self.current_date = jdatetime.date.today()
        self.year_combo.setCurrentText(str(self.current_date.year))
        self.month_combo.setCurrentIndex(self.current_date.month - 1)
    
    def update_calendar(self):
        for i in reversed(range(self.day_grid.count())):
            w = self.day_grid.itemAt(i).widget()
            if w:
                w.deleteLater()
        
        year = int(self.year_combo.currentText())
        month = self.month_combo.currentIndex() + 1
        
        first_day = jdatetime.date(year, month, 1)
        start_weekday = first_day.weekday()
        
        if month <= 6:
            days = 31
        elif month <= 11:
            days = 30
        else:
            days = 30 if first_day.isleap() else 29
        
        row, col = 0, start_weekday
        for day in range(1, days + 1):
            date = jdatetime.date(year, month, day)
            btn = QPushButton(str(day))
            btn.setFixedSize(50, 40)
            if date == jdatetime.date.today():
                btn.setStyleSheet("QPushButton { background: #3b82f6; color: white; border-radius: 10px; font-size: 13px; font-weight: bold; }")
            else:
                btn.setStyleSheet("QPushButton { background: #f1f5f9; color: #1e293b; border-radius: 10px; font-size: 13px; } QPushButton:hover { background: #cbd5e1; }")
            btn.clicked.connect(lambda checked, d=date: self.select_date(d))
            self.day_grid.addWidget(btn, row, col)
            col += 1
            if col > 6:
                col = 0
                row += 1
    
    def select_date(self, date):
        if self.callback:
            self.callback(date.strftime("%Y/%m/%d"))
        self.accept()

# ==================== ویجت تاریخ ====================
class DateSelector(QWidget):
    def __init__(self, label, default_date=None, parent=None):
        super().__init__(parent)
        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(3)
        
        self.label = QLabel(label)
        self.label.setMinimumWidth(50)
        layout.addWidget(self.label)
        
        self.date_edit = QLineEdit()
        self.date_edit.setPlaceholderText("1403/01/01")
        self.date_edit.setFixedWidth(130)
        if default_date:
            self.date_edit.setText(default_date)
        else:
            self.date_edit.setText(jdatetime.date.today().strftime("%Y/%m/%d"))
        layout.addWidget(self.date_edit)
        
        self.calendar_btn = QPushButton("📅")
        self.calendar_btn.setFixedSize(28, 28)
        self.calendar_btn.setStyleSheet("background: #e2e8f0; border-radius: 5px; font-size: 12px;")
        self.calendar_btn.clicked.connect(self.show_calendar)
        layout.addWidget(self.calendar_btn)
    
    def show_calendar(self):
        def on_date_selected(date_str):
            self.date_edit.setText(date_str)
        dialog = PersianCalendarDialog(self, on_date_selected)
        dialog.exec()
    
    def get_date(self):
        return self.date_edit.text()

# ==================== کلاس اصلی ====================
class ReportsTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent = parent
        self.session = SessionLocal()
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        
        header = QLabel("📈 گزارشات حرفه‌ای مرکز مشاوره")
        header.setStyleSheet("font-size: 18px; font-weight: bold; color: #1e293b; padding: 12px; background: linear-gradient(135deg, #e0f2fe, #bae6fd); border-radius: 12px;")
        header.setAlignment(Qt.AlignCenter)
        layout.addWidget(header)
        
        self.inner_tabs = QTabWidget()
        self.inner_tabs.tabBar().setMovable(True)
        
        self.setup_appointments_tab()
        self.setup_finance_tab()
        self.setup_doctors_tab()
        self.setup_patients_tab()
        self.setup_stats_tab()
        
        self.inner_tabs.addTab(self.appointments_tab, "📋 نوبت‌ها")
        self.inner_tabs.addTab(self.finance_tab, "💰 مالی")
        self.inner_tabs.addTab(self.doctors_tab, "👨‍⚕️ اساتید")
        self.inner_tabs.addTab(self.patients_tab, "👥 مراجعین")
        self.inner_tabs.addTab(self.stats_tab, "📊 آمار")
        
        layout.addWidget(self.inner_tabs)
    
    # ========== تب 1: گزارش نوبت‌ها ==========
    def setup_appointments_tab(self):
        self.appointments_tab = QWidget()
        layout = QVBoxLayout(self.appointments_tab)
        
        filter_frame = QFrame()
        filter_frame.setStyleSheet("QFrame { background: #f8fafc; border-radius: 10px; padding: 10px; }")
        filter_layout = QGridLayout(filter_frame)
        filter_layout.setSpacing(10)
        
        self.app_start_date = DateSelector("از تاریخ:", (jdatetime.date.today() - jdatetime.timedelta(days=30)).strftime("%Y/%m/%d"))
        filter_layout.addWidget(self.app_start_date, 0, 0, 1, 2)
        
        self.app_end_date = DateSelector("تا تاریخ:", jdatetime.date.today().strftime("%Y/%m/%d"))
        filter_layout.addWidget(self.app_end_date, 0, 2, 1, 2)
        
        filter_layout.addWidget(QLabel("استاد:"), 1, 0)
        self.app_doctor = QComboBox()
        self.app_doctor.addItem("همه اساتید")
        doctors = self.session.query(Doctor).order_by(Doctor.name).all()
        for doc in doctors:
            self.app_doctor.addItem(doc.name)
        filter_layout.addWidget(self.app_doctor, 1, 1)
        
        filter_layout.addWidget(QLabel("وضعیت:"), 1, 2)
        self.app_status = QComboBox()
        self.app_status.addItems(["همه", "فعال", "انجام شده", "کنسل استاد", "کنسل مراجع"])
        filter_layout.addWidget(self.app_status, 1, 3)
        
        btn_layout = QHBoxLayout()
        self.app_search_btn = QPushButton("🔍 جستجو")
        self.app_search_btn.clicked.connect(self.search_appointments)
        btn_layout.addWidget(self.app_search_btn)
        self.app_excel_btn = QPushButton("📎 اکسل")
        self.app_excel_btn.clicked.connect(self.export_appointments_excel)
        btn_layout.addWidget(self.app_excel_btn)
        self.app_pdf_btn = QPushButton("📄 PDF")
        self.app_pdf_btn.clicked.connect(self.export_appointments_pdf)
        btn_layout.addWidget(self.app_pdf_btn)
        btn_layout.addStretch()
        filter_layout.addLayout(btn_layout, 2, 0, 1, 4)
        
        layout.addWidget(filter_frame)
        
        self.app_table = QTableWidget()
        self.app_table.setAlternatingRowColors(True)
        self.app_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.app_table)
        
        self.app_stats = QLabel("")
        self.app_stats.setStyleSheet("padding: 5px; color: #475569;")
        layout.addWidget(self.app_stats)
    
    def search_appointments(self):
        try:
            start_str = self.app_start_date.get_date()
            end_str = self.app_end_date.get_date()
            start = jdatetime.datetime.strptime(start_str, "%Y/%m/%d").date()
            end = jdatetime.datetime.strptime(end_str, "%Y/%m/%d").date()
        except Exception as e:
            QMessageBox.warning(self, "خطا", f"فرمت تاریخ صحیح نیست. مثال: 1403/01/01")
            return
        
        start_jalali = start.strftime("%Y/%m/%d")
        end_jalali = end.strftime("%Y/%m/%d")
        
        query = self.session.query(Appointment).filter(
            Appointment.date >= start_jalali,
            Appointment.date <= end_jalali
        )
        
        if self.app_doctor.currentText() != "همه اساتید":
            query = query.filter(Appointment.doctor == self.app_doctor.currentText())
        if self.app_status.currentText() != "همه":
            query = query.filter(Appointment.status == self.app_status.currentText())
        
        apps = query.order_by(Appointment.date).all()
        
        headers = ["شناسه", "تاریخ", "ساعت", "استاد", "نام مراجع", "تلفن", "نوع", "وضعیت", "مبلغ"]
        self.app_table.setColumnCount(len(headers))
        self.app_table.setHorizontalHeaderLabels(headers)
        self.app_table.setRowCount(len(apps))
        
        total = 0
        for row, app in enumerate(apps):
            amount = safe_int(app.final_cost)
            total += amount
            self.app_table.setItem(row, 0, QTableWidgetItem(str(app.id)))
            self.app_table.setItem(row, 1, QTableWidgetItem(app.date or ""))
            self.app_table.setItem(row, 2, QTableWidgetItem(app.time or ""))
            self.app_table.setItem(row, 3, QTableWidgetItem(app.doctor or ""))
            self.app_table.setItem(row, 4, QTableWidgetItem(app.patient_name or ""))
            self.app_table.setItem(row, 5, QTableWidgetItem(app.phone or ""))
            self.app_table.setItem(row, 6, QTableWidgetItem(app.type or ""))
            self.app_table.setItem(row, 7, QTableWidgetItem(app.status or ""))
            self.app_table.setItem(row, 8, QTableWidgetItem(f"{amount:,}"))
        
        self.app_stats.setText(f"📊 تعداد: {len(apps)} نوبت | 💰 مجموع: {total:,} تومان")
        self.current_appointments = apps
    
    def export_appointments_excel(self):
        if not hasattr(self, 'current_appointments') or not self.current_appointments:
            QMessageBox.warning(self, "خطا", "لطفاً ابتدا جستجو کنید")
            return
        data = [[a.id, a.date, a.time, a.doctor, a.patient_name, a.phone, a.type, a.status, f"{safe_int(a.final_cost):,}"] for a in self.current_appointments]
        export_to_excel(data, ["شناسه", "تاریخ", "ساعت", "استاد", "نام مراجع", "تلفن", "نوع", "وضعیت", "مبلغ"], "نوبت‌ها", self)
    
    def export_appointments_pdf(self):
        if not hasattr(self, 'current_appointments') or not self.current_appointments:
            QMessageBox.warning(self, "خطا", "لطفاً ابتدا جستجو کنید")
            return
        data = [[a.id, a.date, a.time, a.doctor, a.patient_name, a.phone, a.type, a.status, f"{safe_int(a.final_cost):,}"] for a in self.current_appointments]
        export_to_pdf(data, ["شناسه", "تاریخ", "ساعت", "استاد", "نام مراجع", "تلفن", "نوع", "وضعیت", "مبلغ"], "نوبت‌ها", self)
    
    # ========== تب 2: گزارش مالی ==========
    def setup_finance_tab(self):
        self.finance_tab = QWidget()
        layout = QVBoxLayout(self.finance_tab)
        
        filter_frame = QFrame()
        filter_frame.setStyleSheet("QFrame { background: #f8fafc; border-radius: 10px; padding: 10px; }")
        filter_layout = QGridLayout(filter_frame)
        filter_layout.setSpacing(10)
        
        self.fin_start_date = DateSelector("از تاریخ:", (jdatetime.date.today() - jdatetime.timedelta(days=30)).strftime("%Y/%m/%d"))
        filter_layout.addWidget(self.fin_start_date, 0, 0, 1, 2)
        
        self.fin_end_date = DateSelector("تا تاریخ:", jdatetime.date.today().strftime("%Y/%m/%d"))
        filter_layout.addWidget(self.fin_end_date, 0, 2, 1, 2)
        
        filter_layout.addWidget(QLabel("استاد:"), 1, 0)
        self.fin_doctor = QComboBox()
        self.fin_doctor.addItem("همه اساتید")
        doctors = self.session.query(Doctor).order_by(Doctor.name).all()
        for doc in doctors:
            self.fin_doctor.addItem(doc.name)
        filter_layout.addWidget(self.fin_doctor, 1, 1)
        
        btn_layout = QHBoxLayout()
        self.fin_search_btn = QPushButton("💰 محاسبه")
        self.fin_search_btn.clicked.connect(self.search_finance)
        btn_layout.addWidget(self.fin_search_btn)
        self.fin_excel_btn = QPushButton("📎 اکسل")
        self.fin_excel_btn.clicked.connect(self.export_finance_excel)
        btn_layout.addWidget(self.fin_excel_btn)
        self.fin_pdf_btn = QPushButton("📄 PDF")
        self.fin_pdf_btn.clicked.connect(self.export_finance_pdf)
        btn_layout.addWidget(self.fin_pdf_btn)
        btn_layout.addStretch()
        filter_layout.addLayout(btn_layout, 2, 0, 1, 4)
        
        layout.addWidget(filter_frame)
        
        summary_frame = QFrame()
        summary_frame.setStyleSheet("QFrame { background: #1e40af; border-radius: 12px; padding: 15px; }")
        summary_layout = QHBoxLayout(summary_frame)
        
        self.total_income = QLabel("درآمد کل: 0 تومان")
        self.total_income.setStyleSheet("color: white; font-size: 14px; font-weight: bold;")
        summary_layout.addWidget(self.total_income)
        self.total_doctor = QLabel("سهم اساتید: 0 تومان")
        self.total_doctor.setStyleSheet("color: white; font-size: 14px; font-weight: bold;")
        summary_layout.addWidget(self.total_doctor)
        self.total_center = QLabel("سهم مرکز: 0 تومان")
        self.total_center.setStyleSheet("color: white; font-size: 14px; font-weight: bold;")
        summary_layout.addWidget(self.total_center)
        layout.addWidget(summary_frame)
        
        self.fin_table = QTableWidget()
        self.fin_table.setAlternatingRowColors(True)
        self.fin_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.fin_table)
    
    def search_finance(self):
        try:
            start_str = self.fin_start_date.get_date()
            end_str = self.fin_end_date.get_date()
            start = jdatetime.datetime.strptime(start_str, "%Y/%m/%d").date()
            end = jdatetime.datetime.strptime(end_str, "%Y/%m/%d").date()
        except:
            QMessageBox.warning(self, "خطا", "فرمت تاریخ صحیح نیست. مثال: 1403/01/01")
            return
        
        start_jalali = start.strftime("%Y/%m/%d")
        end_jalali = end.strftime("%Y/%m/%d")
        
        query = self.session.query(Appointment).filter(
            Appointment.date >= start_jalali,
            Appointment.date <= end_jalali,
            Appointment.is_free == 0
        )
        
        if self.fin_doctor.currentText() != "همه اساتید":
            query = query.filter(Appointment.doctor == self.fin_doctor.currentText())
        
        apps = query.all()
        
        total = 0
        doctor_share = 0
        center_share = 0
        
        for app in apps:
            amount = safe_int(app.final_cost)
            total += amount
            try:
                pct = int(app.doc_share.replace("%", "")) if app.doc_share else 60
            except:
                pct = 60
            doctor_share += int(amount * pct / 100)
            center_share += amount - int(amount * pct / 100)
        
        self.total_income.setText(f"درآمد کل: {total:,} تومان")
        self.total_doctor.setText(f"سهم اساتید: {doctor_share:,} تومان")
        self.total_center.setText(f"سهم مرکز: {center_share:,} تومان")
        
        headers = ["شناسه", "تاریخ", "استاد", "نام مراجع", "مبلغ", "درصد استاد", "درصد مرکز", "سهم استاد", "سهم مرکز", "نوع ارجاع"]
        self.fin_table.setColumnCount(len(headers))
        self.fin_table.setHorizontalHeaderLabels(headers)
        self.fin_table.setRowCount(len(apps))
        
        for row, app in enumerate(apps):
            amount = safe_int(app.final_cost)
            try:
                pct = int(app.doc_share.replace("%", "")) if app.doc_share else 60
            except:
                pct = 60
            ref_model = app.ref_model or "مرکز به استاد"
            
            self.fin_table.setItem(row, 0, QTableWidgetItem(str(app.id)))
            self.fin_table.setItem(row, 1, QTableWidgetItem(app.date or ""))
            self.fin_table.setItem(row, 2, QTableWidgetItem(app.doctor or ""))
            self.fin_table.setItem(row, 3, QTableWidgetItem(app.patient_name or ""))
            self.fin_table.setItem(row, 4, QTableWidgetItem(f"{amount:,}"))
            self.fin_table.setItem(row, 5, QTableWidgetItem(f"{pct}%"))
            self.fin_table.setItem(row, 6, QTableWidgetItem(f"{100-pct}%"))
            self.fin_table.setItem(row, 7, QTableWidgetItem(f"{int(amount * pct / 100):,}"))
            self.fin_table.setItem(row, 8, QTableWidgetItem(f"{amount - int(amount * pct / 100):,}"))
            self.fin_table.setItem(row, 9, QTableWidgetItem(ref_model))
        
        self.current_finance = apps
    
    def export_finance_excel(self):
        if not hasattr(self, 'current_finance') or not self.current_finance:
            QMessageBox.warning(self, "خطا", "لطفاً ابتدا محاسبه کنید")
            return
        data = []
        for a in self.current_finance:
            amount = safe_int(a.final_cost)
            try:
                pct = int(a.doc_share.replace("%", "")) if a.doc_share else 60
            except:
                pct = 60
            ref_model = a.ref_model or "مرکز به استاد"
            data.append([
                a.id, a.date, a.doctor, a.patient_name, f"{amount:,}", 
                f"{pct}%", f"{100-pct}%", 
                f"{int(amount * pct / 100):,}", f"{amount - int(amount * pct / 100):,}", 
                ref_model
            ])
        export_to_excel(data, ["شناسه", "تاریخ", "استاد", "نام مراجع", "مبلغ", 
                               "درصد استاد", "درصد مرکز", "سهم استاد", "سهم مرکز", "نوع ارجاع"], 
                        "گزارش_مالی", self)
    
    def export_finance_pdf(self):
        if not hasattr(self, 'current_finance') or not self.current_finance:
            QMessageBox.warning(self, "خطا", "لطفاً ابتدا محاسبه کنید")
            return
        data = []
        for a in self.current_finance:
            amount = safe_int(a.final_cost)
            try:
                pct = int(a.doc_share.replace("%", "")) if a.doc_share else 60
            except:
                pct = 60
            ref_model = a.ref_model or "مرکز به استاد"
            data.append([
                a.id, a.date, a.doctor, a.patient_name, f"{amount:,}", 
                f"{pct}%", f"{100-pct}%", 
                f"{int(amount * pct / 100):,}", f"{amount - int(amount * pct / 100):,}", 
                ref_model
            ])
        export_to_pdf(data, ["شناسه", "تاریخ", "استاد", "نام مراجع", "مبلغ", 
                              "درصد استاد", "درصد مرکز", "سهم استاد", "سهم مرکز", "نوع ارجاع"], 
                       "گزارش_مالی", self)
    
    # ========== تب 3: گزارش اساتید ==========
    def setup_doctors_tab(self):
        self.doctors_tab = QWidget()
        layout = QVBoxLayout(self.doctors_tab)
        
        filter_frame = QFrame()
        filter_frame.setStyleSheet("QFrame { background: #f8fafc; border-radius: 10px; padding: 10px; }")
        filter_layout = QHBoxLayout(filter_frame)
        
        filter_layout.addWidget(QLabel("استاد:"))
        self.doc_doctor = QComboBox()
        self.doc_doctor.addItem("همه اساتید")
        doctors = self.session.query(Doctor).order_by(Doctor.name).all()
        for doc in doctors:
            self.doc_doctor.addItem(doc.name)
        filter_layout.addWidget(self.doc_doctor)
        
        self.doc_search_btn = QPushButton("🔍 نمایش")
        self.doc_search_btn.clicked.connect(self.search_doctors)
        filter_layout.addWidget(self.doc_search_btn)
        self.doc_excel_btn = QPushButton("📎 اکسل")
        self.doc_excel_btn.clicked.connect(self.export_doctors_excel)
        filter_layout.addWidget(self.doc_excel_btn)
        self.doc_pdf_btn = QPushButton("📄 PDF")
        self.doc_pdf_btn.clicked.connect(self.export_doctors_pdf)
        filter_layout.addWidget(self.doc_pdf_btn)
        filter_layout.addStretch()
        layout.addWidget(filter_frame)
        
        self.doc_table = QTableWidget()
        self.doc_table.setAlternatingRowColors(True)
        self.doc_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.doc_table)
    
    def search_doctors(self):
        doctor_name = self.doc_doctor.currentText()
        
        if doctor_name == "همه اساتید":
            doctors = self.session.query(Doctor).all()
        else:
            doctors = [self.session.query(Doctor).filter(Doctor.name == doctor_name).first()]
        
        data = []
        for doc in doctors:
            if not doc:
                continue
            total = self.session.query(Appointment).filter(Appointment.doctor == doc.name).count()
            done = self.session.query(Appointment).filter(Appointment.doctor == doc.name, Appointment.status == "انجام شده").count()
            cancelled = self.session.query(Appointment).filter(
                Appointment.doctor == doc.name,
                Appointment.status.in_(["کنسل استاد", "کنسل مراجع"])
            ).count()
            free = self.session.query(Appointment).filter(Appointment.doctor == doc.name, Appointment.is_free == 1).count()
            apps = self.session.query(Appointment).filter(Appointment.doctor == doc.name, Appointment.is_free == 0).all()
            income = sum(safe_int(a.final_cost) for a in apps)
            data.append([doc.id, doc.name, doc.spec or "-", doc.phone or "-", total, done, cancelled, free, f"{income:,}"])
        
        headers = ["شناسه", "نام", "تخصص", "تلفن", "کل نوبت", "انجام شده", "لغو شده", "رایگان", "درآمد"]
        self.doc_table.setColumnCount(len(headers))
        self.doc_table.setHorizontalHeaderLabels(headers)
        self.doc_table.setRowCount(len(data))
        for row, row_data in enumerate(data):
            for col, val in enumerate(row_data):
                self.doc_table.setItem(row, col, QTableWidgetItem(str(val)))
        self.current_doctors = data
    
    def export_doctors_excel(self):
        if not hasattr(self, 'current_doctors') or not self.current_doctors:
            QMessageBox.warning(self, "خطا", "لطفاً ابتدا جستجو کنید")
            return
        export_to_excel(self.current_doctors, ["شناسه", "نام", "تخصص", "تلفن", "کل نوبت", "انجام شده", "لغو شده", "رایگان", "درآمد"], "اساتید", self)
    
    def export_doctors_pdf(self):
        if not hasattr(self, 'current_doctors') or not self.current_doctors:
            QMessageBox.warning(self, "خطا", "لطفاً ابتدا جستجو کنید")
            return
        export_to_pdf(self.current_doctors, ["شناسه", "نام", "تخصص", "تلفن", "کل نوبت", "انجام شده", "لغو شده", "رایگان", "درآمد"], "اساتید", self)
    
    # ========== تب 4: گزارش مراجعین ==========
    def setup_patients_tab(self):
        self.patients_tab = QWidget()
        layout = QVBoxLayout(self.patients_tab)
        
        filter_frame = QFrame()
        filter_frame.setStyleSheet("QFrame { background: #f8fafc; border-radius: 10px; padding: 10px; }")
        filter_layout = QHBoxLayout(filter_frame)
        
        filter_layout.addWidget(QLabel("جستجو:"))
        self.pat_search = QLineEdit()
        self.pat_search.setPlaceholderText("نام یا کد ملی یا تلفن...")
        self.pat_search.textChanged.connect(self.search_patients)
        filter_layout.addWidget(self.pat_search)
        
        self.pat_excel_btn = QPushButton("📎 اکسل")
        self.pat_excel_btn.clicked.connect(self.export_patients_excel)
        filter_layout.addWidget(self.pat_excel_btn)
        self.pat_pdf_btn = QPushButton("📄 PDF")
        self.pat_pdf_btn.clicked.connect(self.export_patients_pdf)
        filter_layout.addWidget(self.pat_pdf_btn)
        filter_layout.addStretch()
        layout.addWidget(filter_frame)
        
        self.pat_table = QTableWidget()
        self.pat_table.setAlternatingRowColors(True)
        self.pat_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.pat_table)
    
    def search_patients(self):
        search_text = self.pat_search.text().strip()
        
        if search_text:
            patients = self.session.query(Patient).filter(
                (Patient.name.contains(search_text)) |
                (Patient.nat_id.contains(search_text)) |
                (Patient.phone.contains(search_text))
            ).all()
        else:
            patients = self.session.query(Patient).all()
        
        data = []
        for p in patients:
            sessions = self.session.query(Appointment).filter(
                (Appointment.nat_id == p.nat_id) | (Appointment.patient2_nat_id == p.nat_id)
            ).count()
            data.append([p.id, p.name, p.nat_id or "-", p.phone or "-", p.gender or "-", sessions])
        
        headers = ["شناسه", "نام", "کد ملی", "تلفن", "جنسیت", "تعداد جلسات"]
        self.pat_table.setColumnCount(len(headers))
        self.pat_table.setHorizontalHeaderLabels(headers)
        self.pat_table.setRowCount(len(data))
        for row, row_data in enumerate(data):
            for col, val in enumerate(row_data):
                self.pat_table.setItem(row, col, QTableWidgetItem(str(val)))
        self.current_patients = data
    
    def export_patients_excel(self):
        if not hasattr(self, 'current_patients') or not self.current_patients:
            QMessageBox.warning(self, "خطا", "لطفاً ابتدا جستجو کنید")
            return
        export_to_excel(self.current_patients, ["شناسه", "نام", "کد ملی", "تلفن", "جنسیت", "تعداد جلسات"], "مراجعین", self)
    
    def export_patients_pdf(self):
        if not hasattr(self, 'current_patients') or not self.current_patients:
            QMessageBox.warning(self, "خطا", "لطفاً ابتدا جستجو کنید")
            return
        export_to_pdf(self.current_patients, ["شناسه", "نام", "کد ملی", "تلفن", "جنسیت", "تعداد جلسات"], "مراجعین", self)
    
    # ========== تب 5: آمار پیشرفته ==========
    def setup_stats_tab(self):
        self.stats_tab = QWidget()
        layout = QVBoxLayout(self.stats_tab)
        
        filter_frame = QFrame()
        filter_frame.setStyleSheet("QFrame { background: #f8fafc; border-radius: 10px; padding: 10px; }")
        filter_layout = QGridLayout(filter_frame)
        filter_layout.setSpacing(10)
        
        filter_layout.addWidget(QLabel("سال:"), 0, 0)
        self.stats_year = QComboBox()
        for y in range(1400, jdatetime.date.today().year + 2):
            self.stats_year.addItem(str(y))
        self.stats_year.setCurrentText(str(jdatetime.date.today().year))
        filter_layout.addWidget(self.stats_year, 0, 1)
        
        filter_layout.addWidget(QLabel("ماه:"), 0, 2)
        self.stats_month = QComboBox()
        self.stats_month.addItem("کل سال")
        months = ["فروردین", "اردیبهشت", "خرداد", "تیر", "مرداد", "شهریور", 
                  "مهر", "آبان", "آذر", "دی", "بهمن", "اسفند"]
        for m in months:
            self.stats_month.addItem(m)
        filter_layout.addWidget(self.stats_month, 0, 3)
        
        filter_layout.addWidget(QLabel("نوع نمودار:"), 1, 0)
        self.stats_type = QComboBox()
        self.stats_type.addItems(["نوبت به تفکیک استاد", "درآمد ماهیانه", "توزیع نوع نوبت", "توزیع وضعیت نوبت"])
        filter_layout.addWidget(self.stats_type, 1, 1, 1, 3)
        
        self.stats_btn = QPushButton("📊 رسم نمودار")
        self.stats_btn.clicked.connect(self.draw_chart)
        filter_layout.addWidget(self.stats_btn, 2, 0, 1, 2)
        self.stats_save_btn = QPushButton("📎 ذخیره نمودار")
        self.stats_save_btn.clicked.connect(self.save_chart)
        filter_layout.addWidget(self.stats_save_btn, 2, 2, 1, 2)
        
        layout.addWidget(filter_frame)
        
        self.chart_frame = QFrame()
        self.chart_frame.setStyleSheet("QFrame { background: white; border-radius: 10px; }")
        self.chart_layout = QVBoxLayout(self.chart_frame)
        self.chart_label = QLabel("📈 روی دکمه رسم نمودار کلیک کنید")
        self.chart_label.setAlignment(Qt.AlignCenter)
        self.chart_label.setStyleSheet("color: #6b7280; padding: 50px;")
        self.chart_layout.addWidget(self.chart_label)
        
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setWidget(self.chart_frame)
        layout.addWidget(scroll)
        
        self.current_fig = None
    
    def draw_chart(self):
        for i in reversed(range(self.chart_layout.count())):
            w = self.chart_layout.itemAt(i).widget()
            if w:
                w.deleteLater()
        
        year = int(self.stats_year.currentText())
        month = self.stats_month.currentText()
        chart_type = self.stats_type.currentText()
        
        try:
            fig = Figure(figsize=(10, 6), facecolor='white')
            ax = fig.add_subplot(111)
            
            if chart_type == "نوبت به تفکیک استاد":
                if month != "کل سال":
                    month_num = ["فروردین", "اردیبهشت", "خرداد", "تیر", "مرداد", "شهریور", 
                                 "مهر", "آبان", "آذر", "دی", "بهمن", "اسفند"].index(month) + 1
                    month_start = jdatetime.date(year, month_num, 1)
                    if month_num == 12:
                        month_end = jdatetime.date(year + 1, 1, 1) - jdatetime.timedelta(days=1)
                    else:
                        month_end = jdatetime.date(year, month_num + 1, 1) - jdatetime.timedelta(days=1)
                    apps = self.session.query(Appointment).filter(
                        Appointment.date >= month_start.strftime("%Y/%m/%d"),
                        Appointment.date <= month_end.strftime("%Y/%m/%d")
                    ).all()
                else:
                    apps = self.session.query(Appointment).filter(Appointment.date.like(f"{year}/%")).all()
                
                counts = {}
                for app in apps:
                    if app.doctor:
                        counts[app.doctor] = counts.get(app.doctor, 0) + 1
                
                if not counts:
                    self.chart_layout.addWidget(self.chart_label)
                    QMessageBox.warning(self, "اطلاع", "داده‌ای برای نمایش وجود ندارد")
                    return
                
                doctors = list(counts.keys())
                values = list(counts.values())
                y_pos = range(len(doctors))
                ax.barh(y_pos, values, color='#3b82f6')
                ax.set_yticks(y_pos)
                ax.set_yticklabels([persian_text(d) for d in doctors])
                ax.set_xlabel('تعداد نوبت')
                ax.set_title(persian_text(f'نوبت به تفکیک استاد - سال {year}' + (f" - {month}" if month != "کل سال" else "")))
                
            elif chart_type == "درآمد ماهیانه":
                incomes = []
                month_names = ["فروردین", "اردیبهشت", "خرداد", "تیر", "مرداد", "شهریور", 
                              "مهر", "آبان", "آذر", "دی", "بهمن", "اسفند"]
                persian_months = [persian_text(m) for m in month_names]
                
                for m in range(1, 13):
                    m_start = jdatetime.date(year, m, 1)
                    if m == 12:
                        m_end = jdatetime.date(year + 1, 1, 1) - jdatetime.timedelta(days=1)
                    else:
                        m_end = jdatetime.date(year, m + 1, 1) - jdatetime.timedelta(days=1)
                    apps = self.session.query(Appointment).filter(
                        Appointment.date >= m_start.strftime("%Y/%m/%d"),
                        Appointment.date <= m_end.strftime("%Y/%m/%d"),
                        Appointment.is_free == 0
                    ).all()
                    total = sum(safe_int(a.final_cost) for a in apps)
                    incomes.append(total)
                
                ax.bar(persian_months, incomes, color='#10b981')
                ax.set_xlabel(persian_text('ماه'))
                ax.set_ylabel(persian_text('درآمد (تومان)'))
                ax.set_title(persian_text(f'درآمد ماهیانه - سال {year}'))
                plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha='right')
                
            elif chart_type == "توزیع نوع نوبت":
                types = self.session.query(Appointment.type).all()
                counts = {}
                for t in types:
                    name = t[0] or "نامشخص"
                    counts[name] = counts.get(name, 0) + 1
                
                if not counts:
                    self.chart_layout.addWidget(self.chart_label)
                    QMessageBox.warning(self, "اطلاع", "داده‌ای برای نمایش وجود ندارد")
                    return
                
                labels = [persian_text(l) for l in counts.keys()]
                ax.pie(counts.values(), labels=labels, autopct='%1.1f%%', colors=['#3b82f6', '#10b981', '#f59e0b'])
                ax.set_title(persian_text('توزیع نوع نوبت'))
                
            else:
                statuses = self.session.query(Appointment.status).all()
                counts = {}
                for s in statuses:
                    name = s[0] or "نامشخص"
                    counts[name] = counts.get(name, 0) + 1
                
                if not counts:
                    self.chart_layout.addWidget(self.chart_label)
                    QMessageBox.warning(self, "اطلاع", "داده‌ای برای نمایش وجود ندارد")
                    return
                
                labels = [persian_text(l) for l in counts.keys()]
                ax.pie(counts.values(), labels=labels, autopct='%1.1f%%', 
                       colors=['#3b82f6', '#10b981', '#ef4444', '#f59e0b', '#8b5cf6'])
                ax.set_title(persian_text('توزیع وضعیت نوبت'))
            
            fig.tight_layout()
            canvas = FigureCanvas(fig)
            self.current_fig = fig
            self.chart_layout.addWidget(canvas)
        except Exception as e:
            self.chart_layout.addWidget(self.chart_label)
            QMessageBox.critical(self, "خطا", f"خطا: {str(e)}")
    
    def save_chart(self):
        if self.current_fig is None:
            QMessageBox.warning(self, "خطا", "لطفاً ابتدا نموداری رسم کنید")
            return
        filepath, _ = QFileDialog.getSaveFileName(self, "ذخیره نمودار", "chart.png", "PNG Image (*.png)")
        if filepath:
            self.current_fig.savefig(filepath, dpi=300, bbox_inches='tight')
            QMessageBox.information(self, "موفق", f"نمودار ذخیره شد:\n{filepath}")
    
    def __del__(self):
        if hasattr(self, 'session'):
            try:
                self.session.close()
            except:
                pass





